import uuid
from decimal import Decimal
from django.db import models
from django.db.models import Count, Q, Sum
from django.contrib.auth import get_user_model
from django.utils.text import slugify
from django.utils import timezone

from rest_framework import viewsets, permissions, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
import django_filters as filters_filter

class CharInFilter(filters_filter.BaseInFilter, filters_filter.CharFilter):
    pass


from project_management.models import (
    Workspace, Project, ProjectMember, Workflow,
    WorkItemStatus, WorkItem, WorkItemLink, WorkItemComment,
    WorkItemAttachment, WorkItemActivityLog, log_activity,
    Sprint, SprintMember, Release, SLAPolicy, CSATResponse, Milestone,
    Objective, KeyResult, AutomationRule,
    Notification, NotificationPreference, WebhookConfig,
    CustomFieldDefinition, ProjectFieldVisibility,
    RecurringTaskConfig, ApprovalWorkflow, ApprovalStep,
    ApprovalRequest, ApprovalAction, CannedResponse,
    ProjectTemplate, GitHubIntegration, GitHubLink, PriorityRule,
    WorkItemTimeLog, WebhookDeliveryLog, TransitionRule,
    SavedFilter, WorkItemTemplate,
)
from project_management.serializers import (
    WorkspaceSerializer, ProjectSerializer, ProjectListSerializer,
    ProjectMemberSerializer, WorkflowSerializer, WorkItemStatusSerializer,
    WorkItemSerializer, WorkItemDetailSerializer,
    WorkItemLinkSerializer, WorkItemCommentSerializer,
    WorkItemAttachmentSerializer, WorkItemActivityLogSerializer, UserBriefSerializer,
    SprintSerializer, SprintDetailSerializer, SprintMemberSerializer,
    ReleaseSerializer, SLAPolicySerializer, CSATResponseSerializer, MilestoneSerializer,
    ObjectiveListSerializer, ObjectiveDetailSerializer, KeyResultSerializer,
    AutomationRuleSerializer,
    NotificationSerializer, NotificationPreferenceSerializer, WebhookConfigSerializer,
    CustomFieldDefinitionSerializer, ProjectFieldVisibilitySerializer,
    RecurringTaskConfigSerializer, ApprovalWorkflowSerializer,
    ApprovalStepSerializer, ApprovalRequestSerializer, ApprovalActionSerializer,
    CannedResponseSerializer, ProjectTemplateSerializer,
    GitHubIntegrationSerializer, GitHubLinkSerializer, PriorityRuleSerializer,
    WebhookDeliveryLogSerializer, TransitionRuleSerializer,
    SavedFilterSerializer, WorkItemTemplateSerializer,
)
from project_management.services.sla_engine import SLAEngine

try:
    from sales_task_manager.models import SalesTarget, TargetCycle
    SALES_TASK_MANAGER_AVAILABLE = True
except ImportError:
    SalesTarget = None
    TargetCycle = None
    SALES_TASK_MANAGER_AVAILABLE = False
from project_management.permissions import (
    IsOrgAdmin, IsWorkspaceAdmin, IsProjectAdminOrEditor,
    IsProjectMember, IsProjectViewer, CanManageWorkflows
)
from project_management.services.burndown_engine import BurndownEngine

User = get_user_model()


# ============================================================================
# WORKSPACE VIEWSET
# ============================================================================

class WorkspaceViewSet(viewsets.ModelViewSet):
    queryset = Workspace.objects.all().prefetch_related('projects')
    serializer_class = WorkspaceSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    search_fields = ['name', 'description']

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsOrgAdmin()]
        return [permissions.IsAuthenticated()]

    @action(detail=True, methods=['get'])
    def summary(self, request, pk=None):
        workspace = self.get_object()
        projects = workspace.projects.all()
        total_items = WorkItem.objects.filter(project__in=projects)
        active_sprints = Sprint.objects.filter(project__in=projects, status='ACTIVE').count()

        return Response({
            'id': str(workspace.id),
            'name': workspace.name,
            'slug': workspace.slug,
            'project_count': projects.count(),
            'active_projects': projects.filter(is_active=True).count(),
            'total_work_items': total_items.count(),
            'open_items': total_items.exclude(status__category='done').count(),
            'overdue_items': total_items.filter(
                due_date__lt=timezone.now(),
                status__category__in=['todo', 'in_progress']
            ).count(),
            'active_sprints': active_sprints,
        })


# ============================================================================
# PROJECT VIEWSET
# ============================================================================

class ProjectViewSet(viewsets.ModelViewSet):
    queryset = Project.objects.all().select_related(
        'workspace', 'department', 'parent_project'
    ).prefetch_related('members', 'members__user', 'workflows', 'workflows__statuses')
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['workspace', 'department', 'is_active']
    search_fields = ['name', 'key', 'description']

    def get_serializer_class(self):
        if self.action == 'list':
            return ProjectListSerializer
        return ProjectSerializer

    def perform_create(self, serializer):
        preset = self.request.data.get('workflow_preset')
        project = serializer.save()
        if preset:
            from project_management.services.workflow_engine import apply_workflow_preset
            apply_workflow_preset(project, preset)
        return project

    def get_permissions(self):
        if self.action in ['create', 'destroy']:
            return [IsWorkspaceAdmin()]
        elif self.action in ['update', 'partial_update', 'add_member', 'remove_member']:
            return [IsProjectAdminOrEditor()]
        return [permissions.IsAuthenticated()]

    def get_queryset(self):
        user = self.request.user
        qs = super().get_queryset()
        if user.is_authenticated and user.role == 'Staff':
            qs = qs.filter(
                Q(department__in=user.departments.all()) |
                Q(members__user=user)
            ).distinct()
        return qs

    @action(detail=True, methods=['get'])
    def board(self, request, pk=None):
        """Get work items organized by status for Kanban board view."""
        project = Project.objects.filter(id=pk, is_active=True).first() or self.get_object()
        sprint_id = request.query_params.get('sprint')
        statuses = self._project_statuses(project)

        items_qs = WorkItem.objects.filter(project=project)
        if sprint_id:
            items_qs = items_qs.filter(sprint_id=sprint_id)

        items_qs = items_qs.select_related(
            'assignee', 'reporter', 'epic', 'sprint', 'status'
        ).annotate(
            subtask_count=models.Count('subtasks', distinct=True),
            comment_count=models.Count('comments', distinct=True),
        ).order_by('order')

        grouped_items = {}
        all_items = list(items_qs)
        for item in all_items:
            grouped_items.setdefault(str(item.status_id), []).append(item)

        columns = []
        for status_obj in statuses:
            column_items = grouped_items.get(str(status_obj.id), [])
            columns.append({
                'id': str(status_obj.id),
                'name': status_obj.name,
                'slug': status_obj.slug,
                'color': status_obj.color,
                'category': status_obj.category,
                'order': status_obj.order,
                'items': WorkItemSerializer(column_items, many=True).data,
                'item_count': len(column_items),
            })

        # Cache counts so ProjectListSerializer doesn't re-query them.
        project.member_count_annotated = project.members.count()
        project.work_item_summary_annotated = {
            'total': len(all_items),
            'todo': sum(1 for i in all_items if i.status.category == 'todo'),
            'in_progress': sum(1 for i in all_items if i.status.category == 'in_progress'),
            'done': sum(1 for i in all_items if i.status.category == 'done'),
        }

        return Response({
            'project': ProjectListSerializer(project).data,
            'columns': columns,
            'active_sprints': SprintSerializer(
                Sprint.objects.filter(project=project, status='ACTIVE'), many=True
            ).data,
        })

    @staticmethod
    def _project_statuses(project):
        """Return the statuses to display for a project's board.

        Uses statuses mapped to the project's workflows; if the project has no
        workflow mapped yet (so the board would otherwise be empty), fall back
        to the default dev-kanban workflow's statuses so columns always show.
        """
        statuses = list(WorkItemStatus.objects.filter(
            workflow__projects=project
        ).order_by('workflow', 'order'))
        if not statuses:
            default_wf = Workflow.objects.filter(
                is_default=True, scope='dev_kanban'
            ).order_by('id').first()
            if default_wf:
                statuses = list(WorkItemStatus.objects.filter(
                    workflow=default_wf
                ).order_by('order'))
        return statuses

    @action(detail=True, methods=['get'])
    def backlog(self, request, pk=None):
        """Backlog view — unsorted/unassigned items not in active sprint."""
        project = self.get_object()
        active_sprint = Sprint.objects.filter(project=project, status='ACTIVE').first()

        backlog_items = WorkItem.objects.filter(
            project=project,
            status__category__in=['backlog', 'todo']
        )

        # Exclude items already in active sprint
        if active_sprint:
            backlog_items = backlog_items.exclude(sprint=active_sprint)

        backlog_items = backlog_items.select_related('assignee', 'epic', 'sprint'
        ).order_by('order', 'priority')

        # Group by epic
        epics = WorkItem.objects.filter(
            project=project, issue_type='EPIC'
        ).values('id', 'key', 'title')

        grouped = []
        for epic in epics:
            epic_items = backlog_items.filter(epic_id=epic['id'])
            if epic_items.exists():
                grouped.append({
                    'epic': {'id': str(epic['id']), 'key': epic['key'], 'title': epic['title']},
                    'items': WorkItemSerializer(list(epic_items), many=True).data,
                })
                backlog_items = backlog_items.exclude(epic_id=epic['id'])

        # Unassigned to any epic
        if backlog_items.exists():
            grouped.append({
                'epic': None,
                'items': WorkItemSerializer(list(backlog_items), many=True).data,
            })

        return Response({
            'project': ProjectListSerializer(project).data,
            'active_sprint': SprintSerializer(active_sprint).data if active_sprint else None,
            'backlog': grouped,
            'summary': {
                'total_items': sum(len(g['items']) for g in grouped),
                'total_points': backlog_items.aggregate(total=Sum('story_points'))['total'] or 0,
            },
        })

    @action(detail=True, methods=['get'])
    def velocity(self, request, pk=None):
        """Team velocity from last N sprints."""
        project = self.get_object()
        n = int(request.query_params.get('sprints', 5))
        velocity = BurndownEngine.calculate_velocity(project, n)
        return Response(velocity)

    @action(detail=True, methods=['get'])
    def epics(self, request, pk=None):
        """Get all epics for this project with completion stats."""
        project = self.get_object()
        epics = WorkItem.objects.filter(project=project, issue_type='EPIC')

        epic_data = []
        for epic in epics:
            children = WorkItem.objects.filter(epic=epic)
            total_points = children.aggregate(total=Sum('story_points'))['total'] or 0
            done_points = children.filter(
                status__category='done'
            ).aggregate(total=Sum('story_points'))['total'] or 0

            epic_data.append({
                'id': str(epic.id),
                'key': epic.key,
                'title': epic.title,
                'status': WorkItemStatusSerializer(epic.status).data,
                'priority': epic.priority,
                'assignee': UserBriefSerializer(epic.assignee).data if epic.assignee else None,
                'child_count': children.count(),
                'total_points': float(total_points),
                'completed_points': float(done_points),
                'completion_pct': round(
                    (float(done_points) / float(total_points) * 100) if total_points else 0, 1
                ),
                'due_date': epic.due_date.isoformat() if epic.due_date else None,
            })

        return Response(epic_data)

    @action(detail=True, methods=['post'], url_path='add-member')
    def add_member(self, request, pk=None):
        project = self.get_object()
        user_id = request.data.get('user_id')
        role = request.data.get('role', 'MEMBER')

        if not user_id:
            return Response({'error': 'user_id is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)

        member, created = ProjectMember.objects.get_or_create(
            project=project, user=user, defaults={'role': role}
        )
        if not created:
            member.role = role
            member.save()

        return Response(ProjectMemberSerializer(member).data)

    @action(detail=True, methods=['post'], url_path='remove-member')
    def remove_member(self, request, pk=None):
        project = self.get_object()
        user_id = request.data.get('user_id')
        if not user_id:
            return Response({'error': 'user_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        deleted, _ = ProjectMember.objects.filter(project=project, user_id=user_id).delete()
        return Response({'success': True, 'removed': deleted > 0})

    # ═══════════════════════════════════════════════════════════════════
    # Phase 2: Sales Pipeline View
    # ═══════════════════════════════════════════════════════════════════

    @action(detail=True, methods=['get'])
    def pipeline(self, request, pk=None):
        """Sales pipeline view — WorkItems with issue_type=DEAL organized by status columns.
        Includes deal_value, probability from custom_fields.
        """
        project = self.get_object()
        assignee = request.query_params.get('assignee')
        stage = request.query_params.get('stage')

        statuses = WorkItemStatus.objects.filter(
            workflow__projects=project
        ).order_by('workflow', 'order')

        items_qs = WorkItem.objects.filter(project=project, issue_type='DEAL')
        if assignee:
            items_qs = items_qs.filter(assignee_id=assignee)
        if stage:
            items_qs = items_qs.filter(status_id=stage)

        columns = []
        for status_obj in statuses:
            items = items_qs.filter(status=status_obj).select_related(
                'assignee', 'reporter', 'epic', 'sprint'
            ).order_by('order')

            serialized_items = WorkItemSerializer(items, many=True).data
            # Enrich with deal-specific computed fields
            enriched = []
            for item_data, item_obj in zip(serialized_items, items):
                cf = item_obj.custom_fields or {}
                item_data['deal_value'] = cf.get('deal_value', 0)
                item_data['probability'] = cf.get('probability', 0)
                item_data['weighted_value'] = round(
                    float(cf.get('deal_value', 0)) * float(cf.get('probability', 0)) / 100, 2
                )
                item_data['expected_close_date'] = cf.get('expected_close_date')
                enriched.append(item_data)

            stage_total = sum(
                float(i.get('deal_value', 0)) for i in enriched
            )
            stage_weighted = sum(
                float(i.get('weighted_value', 0)) for i in enriched
            )

            columns.append({
                'id': str(status_obj.id),
                'name': status_obj.name,
                'slug': status_obj.slug,
                'color': status_obj.color,
                'category': status_obj.category,
                'order': status_obj.order,
                'items': enriched,
                'item_count': len(enriched),
                'stage_total': stage_total,
                'stage_weighted': stage_weighted,
            })

        total_pipeline_value = sum(c['stage_total'] for c in columns)
        total_weighted_forecast = sum(c['stage_weighted'] for c in columns)

        return Response({
            'project': ProjectListSerializer(project).data,
            'columns': columns,
            'summary': {
                'total_deals': sum(c['item_count'] for c in columns),
                'total_pipeline_value': total_pipeline_value,
                'total_weighted_forecast': total_weighted_forecast,
            },
        })

    # ═══════════════════════════════════════════════════════════════════
    # Phase 2.5: Ticket Queue
    # ═══════════════════════════════════════════════════════════════════

    @action(detail=True, methods=['get'])
    def ticket_queue(self, request, pk=None):
        """Support ticket queue sorted by SLA breach risk, priority, and age."""
        project = self.get_object()
        sort_by = request.query_params.get('sort', 'sla')
        assignee = request.query_params.get('assignee')
        status_filter = request.query_params.get('status')

        tickets = WorkItem.objects.filter(
            project=project, issue_type='TICKET'
        ).select_related('assignee', 'reporter', 'status', 'sla_policy')

        if assignee:
            tickets = tickets.filter(assignee_id=assignee)
        if status_filter:
            tickets = tickets.filter(status_id=status_filter)

        # Annotate with SLA breaches for sorting
        ticket_data = []
        for t in tickets:
            sla = SLAEngine.check_sla(t)
            ticket_data.append({
                'id': str(t.id),
                'key': t.key,
                'title': t.title,
                'status': WorkItemStatusSerializer(t.status).data,
                'priority': t.priority,
                'assignee': UserBriefSerializer(t.assignee).data if t.assignee else None,
                'reporter': UserBriefSerializer(t.reporter).data if t.reporter else None,
                'sla_status': sla['status'],
                'sla_breached': sla['breached'],
                'response_remaining_minutes': sla.get('response_remaining_minutes'),
                'resolution_remaining_minutes': sla.get('resolution_remaining_minutes'),
                'first_response_at': t.first_response_at.isoformat() if t.first_response_at else None,
                'requester_name': t.requester_name,
                'requester_email': t.requester_email,
                'comment_count': t.comments.count(),
                'created_at': t.created_at.isoformat(),
                'age_hours': round((timezone.now() - t.created_at).total_seconds() / 3600, 1),
            })

        # Sort tickets
        priority_order = {'CRITICAL': 0, 'HIGH': 1, 'MEDIUM': 2, 'LOW': 3}
        if sort_by == 'sla':
            ticket_data.sort(key=lambda x: (
                0 if x['sla_breached'] else (1 if x['sla_status'] == 'WARNING' else 2),
                x.get('response_remaining_minutes') or x.get('resolution_remaining_minutes') or 9999,
                priority_order.get(x['priority'], 99),
            ))
        elif sort_by == 'priority':
            ticket_data.sort(key=lambda x: (priority_order.get(x['priority'], 99), -x['age_hours']))
        elif sort_by == 'age':
            ticket_data.sort(key=lambda x: -x['age_hours'])

        # Summary stats
        total = len(ticket_data)
        breached = sum(1 for t in ticket_data if t['sla_breached'])
        warning = sum(1 for t in ticket_data if t['sla_status'] == 'WARNING')
        unassigned = sum(1 for t in ticket_data if not t['assignee'])

        return Response({
            'tickets': ticket_data,
            'summary': {
                'total': total,
                'breached': breached,
                'warning': warning,
                'unassigned': unassigned,
                'open': total - sum(1 for t in ticket_data if t['status']['category'] == 'done'),
            },
        })

    # ═══════════════════════════════════════════════════════════════════
    # Phase 3: Team sync + Hierarchy
    # ═══════════════════════════════════════════════════════════════════

    @action(detail=True, methods=['post'])
    def sync_members(self, request, pk=None):
        """Sync project members from the linked HR department."""
        project = self.get_object()
        if not project.department:
            return Response({'error': 'No HR department linked to this project'}, status=400)

        from hr.models import Employee
        employees = Employee.objects.filter(
            department=project.department,
            status='ACTIVE',
        ).select_related('user')

        added = []
        for emp in employees:
            if emp.user and not project.members.filter(user=emp.user).exists():
                ProjectMember.objects.create(
                    project=project,
                    user=emp.user,
                    role='MEMBER',
                )
                added.append({'id': str(emp.user.id), 'name': emp.user.get_full_name() or emp.user.email})

        return Response({
            'synced': True,
            'members_added': len(added),
            'total_members': project.members.count(),
            'added': added,
        })

    @action(detail=True, methods=['get'])
    def hierarchy(self, request, pk=None):
        """Get the full project hierarchy tree."""
        project = self.get_object()
        ancestors = []
        current = project
        while current.parent_project:
            current = current.parent_project
            ancestors.append({
                'id': str(current.id),
                'name': current.name,
                'key': current.key,
            })
        descendants = []
        for sub in Project.objects.filter(parent_project=project).select_related('parent_project'):
            descendants.append({
                'id': str(sub.id),
                'name': sub.name,
                'key': sub.key,
                'work_item_count': sub.work_items.count(),
            })
        return Response({
            'project': {'id': str(project.id), 'name': project.name, 'key': project.key},
            'ancestors': list(reversed(ancestors)),
            'descendants': descendants,
        })


# ============================================================================
# SPRINT VIEWSET — Phase 1: Dev Mode
# ============================================================================

class SprintViewSet(viewsets.ModelViewSet):
    queryset = Sprint.objects.all().select_related('project').prefetch_related('members')
    serializer_class = SprintSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['project', 'status']
    search_fields = ['name', 'goal']

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return SprintDetailSerializer
        return SprintSerializer

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy', 'start', 'close']:
            return [IsProjectAdminOrEditor()]
        return [permissions.IsAuthenticated()]

    @action(detail=True, methods=['post'])
    def start(self, request, pk=None):
        """Start a sprint: set status to ACTIVE and log activity."""
        sprint = self.get_object()
        if sprint.status != 'PLANNING':
            return Response(
                {'error': f'Cannot start sprint in {sprint.get_status_display()} status'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Close any other active sprint in the same project
        Sprint.objects.filter(project=sprint.project, status='ACTIVE').update(status='COMPLETED')

        sprint.status = 'ACTIVE'
        sprint.save()

        log_activity(
            work_item=None,
            user=request.user,
            activity_type='SPRINT_STARTED',
            description=f"Sprint '{sprint.name}' started",
            metadata={'sprint_id': str(sprint.id), 'project_id': str(sprint.project_id)}
        )

        return Response(SprintDetailSerializer(sprint).data)

    @action(detail=True, methods=['post'])
    def close(self, request, pk=None):
        """Close a sprint: calculate final metrics, move incomplete items back to backlog."""
        sprint = self.get_object()
        if sprint.status != 'ACTIVE':
            return Response(
                {'error': f'Cannot close sprint in {sprint.get_status_display()} status'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Calculate final stats
        sprint_items = WorkItem.objects.filter(sprint=sprint)
        sprint.total_committed_points = sprint_items.aggregate(
            total=Sum('story_points')
        )['total'] or 0

        completed = sprint_items.filter(status__category='done')
        sprint.total_completed_points = completed.aggregate(
            total=Sum('story_points')
        )['total'] or 0

        sprint.status = 'COMPLETED'
        sprint.completed_at = timezone.now()
        sprint.save()

        # Move incomplete items back to backlog (remove from sprint)
        incomplete = sprint_items.exclude(status__category='done')
        count = incomplete.count()
        incomplete.update(sprint=None)

        log_activity(
            work_item=None,
            user=request.user,
            activity_type='SPRINT_CLOSED',
            description=f"Sprint '{sprint.name}' closed. {count} incomplete items moved to backlog.",
            metadata={
                'sprint_id': str(sprint.id),
                'committed_points': float(sprint.total_committed_points),
                'completed_points': float(sprint.total_completed_points),
                'moved_to_backlog': count,
            }
        )

        stats = BurndownEngine.get_sprint_stats(sprint)
        return Response({
            'sprint': SprintDetailSerializer(sprint).data,
            'stats': stats,
            'items_moved_to_backlog': count,
        })

    @action(detail=True, methods=['get'])
    def burndown(self, request, pk=None):
        """Get burndown chart data for a sprint."""
        sprint = self.get_object()
        data = BurndownEngine.calculate_burndown(sprint)
        return Response(data)

    @action(detail=True, methods=['get'])
    def stats(self, request, pk=None):
        """Comprehensive sprint statistics."""
        sprint = self.get_object()
        stats = BurndownEngine.get_sprint_stats(sprint)

        # Velocity context
        velocity = BurndownEngine.calculate_velocity(sprint.project)
        stats['velocity_context'] = velocity

        return Response(stats)

    @action(detail=False, methods=['get'])
    def recommendations(self, request):
        """AI-powered sprint recommendations based on velocity and backlog."""
        from project_management.services.recommendation_engine import get_sprint_recommendations
        project_id = request.query_params.get('project')
        sprint_id = request.query_params.get('sprint')
        if not project_id:
            return Response({'error': 'project parameter is required'}, status=400)
        data = get_sprint_recommendations(project_id=project_id, sprint_id=sprint_id)
        return Response(data)

    @action(detail=True, methods=['post'], url_path='add-member')
    def add_member(self, request, pk=None):
        """Add a team member to this sprint with capacity."""
        sprint = self.get_object()
        user_id = request.data.get('user_id')
        capacity = request.data.get('capacity_hours', 40)

        if not user_id:
            return Response({'error': 'user_id is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)

        member, created = SprintMember.objects.get_or_create(
            sprint=sprint, user=user, defaults={'capacity_hours': capacity}
        )
        if not created:
            member.capacity_hours = capacity
            member.save()

        return Response(SprintMemberSerializer(member).data)

    @action(detail=True, methods=['post'], url_path='remove-member')
    def remove_member(self, request, pk=None):
        """Remove a team member from this sprint."""
        sprint = self.get_object()
        user_id = request.data.get('user_id')
        if not user_id:
            return Response({'error': 'user_id is required'}, status=status.HTTP_400_BAD_REQUEST)

        deleted, _ = SprintMember.objects.filter(
            sprint=sprint, user_id=user_id
        ).delete()
        return Response({'success': True, 'removed': deleted > 0})


# ============================================================================
# RELEASE VIEWSET — Phase 1: Dev Mode
# ============================================================================

class ReleaseViewSet(viewsets.ModelViewSet):
    queryset = Release.objects.all().select_related('project')
    serializer_class = ReleaseSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['project', 'status', 'is_archived']
    search_fields = ['name', 'description', 'version']

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsProjectAdminOrEditor()]
        return [permissions.IsAuthenticated()]


# ============================================================================
# WORKFLOW VIEWSET
# ============================================================================

class WorkflowViewSet(viewsets.ModelViewSet):
    queryset = Workflow.objects.all().prefetch_related('statuses', 'projects')
    serializer_class = WorkflowSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['scope', 'is_default']
    search_fields = ['name', 'description']

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [CanManageWorkflows()]
        return [permissions.IsAuthenticated()]

    @action(detail=True, methods=['post'], url_path='add-status')
    def add_status(self, request, pk=None):
        workflow = self.get_object()
        name = request.data.get('name')
        color = request.data.get('color', 'gray')
        category = request.data.get('category', 'todo')

        if not name:
            return Response({'error': 'name is required'}, status=status.HTTP_400_BAD_REQUEST)

        base_slug = slugify(name)
        if not base_slug:
            base_slug = f'status-{uuid.uuid4().hex[:8]}'

        existing_slugs = WorkItemStatus.objects.filter(
            workflow=workflow
        ).values_list('slug', flat=True)
        test_slug = base_slug
        counter = 1
        while test_slug in existing_slugs:
            test_slug = f"{base_slug}-{counter}"
            counter += 1

        max_order = WorkItemStatus.objects.filter(
            workflow=workflow
        ).aggregate(models.Max('order'))['order__max'] or 0

        status_obj = WorkItemStatus.objects.create(
            workflow=workflow, name=name, slug=test_slug,
            order=request.data.get('order', max_order + 1),
            color=color, category=category,
            is_start=request.data.get('is_start', False),
            is_end=request.data.get('is_end', False),
        )
        return Response(WorkItemStatusSerializer(status_obj).data)

    @action(detail=True, methods=['post'], url_path='reorder-statuses')
    def reorder_statuses(self, request, pk=None):
        items = request.data
        for item in items:
            WorkItemStatus.objects.filter(id=item['id'], workflow_id=pk).update(order=item['order'])
        return Response({'success': True})


# ============================================================================
# WORK ITEM VIEWSET
# ============================================================================

class WorkItemFilterSet(filters_filter.FilterSet):
    """Explicit filterset so status fields (e.g. status__category, __isnull)
    work instead of returning 400 Bad Request."""
    project = filters_filter.UUIDFilter(field_name='project_id')
    epic = filters_filter.UUIDFilter(field_name='epic_id')
    parent = filters_filter.UUIDFilter(field_name='parent_id')
    sprint = filters_filter.UUIDFilter(field_name='sprint_id')
    assignee = filters_filter.UUIDFilter(field_name='assignee_id')
    status = filters_filter.UUIDFilter(field_name='status_id')
    status__category = filters_filter.CharFilter(field_name='status__category')
    status__category__in = CharInFilter(field_name='status__category', lookup_expr='in')
    sprint__isnull = filters_filter.BooleanFilter(method='filter_isnull')
    epic__isnull = filters_filter.BooleanFilter(method='filter_isnull')
    parent__isnull = filters_filter.BooleanFilter(method='filter_isnull')

    def filter_isnull(self, queryset, name, value):
        field_name = name.replace('__isnull', '')
        if value is not None:
            return queryset.filter(**{f'{field_name}__isnull': value})
        return queryset
    issue_type = filters_filter.CharFilter(field_name='issue_type')
    priority = filters_filter.CharFilter(field_name='priority')
    due_date__gte = filters_filter.DateTimeFilter(field_name='due_date', lookup_expr='gte')
    due_date__lte = filters_filter.DateTimeFilter(field_name='due_date', lookup_expr='lte')
    story_points__gte = filters_filter.NumberFilter(field_name='story_points', lookup_expr='gte')
    story_points__lte = filters_filter.NumberFilter(field_name='story_points', lookup_expr='lte')

    class Meta:
        model = WorkItem
        fields = [
            'project', 'status', 'assignee', 'issue_type', 'priority',
            'parent', 'epic', 'sprint',
        ]


class WorkItemViewSet(viewsets.ModelViewSet):
    queryset = WorkItem.objects.all().select_related(
        'project', 'status', 'assignee', 'reporter', 'epic', 'sprint'
    ).prefetch_related(
        'subtasks', 'comments', 'watchers',
        'outgoing_links', 'incoming_links',
        'activity_logs'
    ).annotate(
        subtask_count=models.Count('subtasks', distinct=True),
        comment_count=models.Count('comments', distinct=True),
    )
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_class = WorkItemFilterSet
    search_fields = ['title', 'description', 'key']

    def list(self, request, *args, **kwargs):
        """Override list to convert string 'null' → __isnull filter params."""
        null_keys = ['parent', 'epic', 'sprint', 'assignee', 'status', 'project']
        for key in null_keys:
            val = request.query_params.get(key)
            if val and val.strip().lower() == 'null':
                self._fix_null_param(request, key)
        return super().list(request, *args, **kwargs)

    def _fix_null_param(self, request, key):
        """Replace 'field=null' with 'field__isnull=true'."""
        try:
            qd = request._request.GET
            if hasattr(qd, '_mutable'):
                qd._mutable = True
            if key in qd:
                qd[f'{key}__isnull'] = 'true'
                del qd[key]
        except Exception:
            pass

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return WorkItemDetailSerializer
        return WorkItemSerializer

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update']:
            return [IsProjectAdminOrEditor()]
        elif self.action == 'destroy':
            return [IsProjectAdminOrEditor()]
        return [permissions.IsAuthenticated()]

    def get_queryset(self):
        user = self.request.user
        qs = super().get_queryset()
        if user.is_authenticated and user.role == 'Staff':
            qs = qs.filter(
                Q(assignee=user) | Q(project__members__user=user)
            ).distinct()
        return qs

    def create(self, request, *args, **kwargs):
        data = request.data.copy() if hasattr(request.data, 'copy') else {**request.data}
        if not data.get('status'):
            from project_management.models import WorkItemStatus
            project_id = data.get('project')
            first_status = None
            if project_id:
                first_status = WorkItemStatus.objects.filter(
                    workflow__projects=project_id
                ).order_by(
                    'order', 'is_start'
                ).first()
            if not first_status:
                # Fall back to the default dev-kanban workflow's start status so
                # created items land in a visible board column even for projects
                # that have no workflow mapped yet.
                default_wf = Workflow.objects.filter(
                    is_default=True, scope='dev_kanban'
                ).order_by('id').first()
                if default_wf:
                    first_status = (
                        WorkItemStatus.objects.filter(
                            workflow=default_wf, is_start=True
                        ).order_by('order').first()
                        or WorkItemStatus.objects.filter(
                            workflow=default_wf
                        ).order_by('order').first()
                    )
            if not first_status:
                # Last resort: any start status in the system.
                first_status = (
                    WorkItemStatus.objects.filter(is_start=True).order_by('order').first()
                    or WorkItemStatus.objects.order_by('order').first()
                )
            if first_status:
                data['status'] = str(first_status.id)
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def perform_create(self, serializer):
        work_item = serializer.save(reporter=self.request.user)

        # Phase 2.5: Auto-start SLA for ticket items
        if work_item.issue_type == 'TICKET':
            self._auto_assign_ticket(work_item)
            if work_item.sla_policy:
                SLAEngine.start_sla(work_item)

        log_activity(
            work_item=work_item, user=self.request.user,
            activity_type='ITEM_CREATED',
            description=f"{work_item.get_issue_type_display()} '{work_item.title}' created",
            metadata={'issue_type': work_item.issue_type, 'key': work_item.key}
        )

    def _auto_assign_ticket(self, work_item):
        """Simple round-robin auto-assignment for unassigned tickets."""
        if work_item.assignee:
            return
        project_members = ProjectMember.objects.filter(
            project=work_item.project,
            role__in=['ADMIN', 'EDITOR', 'MEMBER']
        ).select_related('user').order_by('user__id')

        if not project_members.exists():
            return

        # Round-robin: pick the member with fewest open tickets
        from django.db.models import Count
        member_loads = {
            m.user_id: WorkItem.objects.filter(
                assignee=m.user,
                project=work_item.project,
                status__category__in=['todo', 'in_progress']
            ).count()
            for m in project_members
        }

        best_member = min(project_members, key=lambda m: member_loads.get(m.user_id, 0))
        work_item.assignee = best_member.user

        log_activity(
            work_item=work_item, user=None,
            activity_type='TICKET_ASSIGNED',
            description=f"Auto-assigned to {best_member.user.get_full_name()}",
            metadata={'assignee_id': str(best_member.user_id)}
        )

    def perform_update(self, serializer):
        old = self.get_object()
        work_item = serializer.save()

        if old.status_id != work_item.status_id:
            log_activity(
                work_item=work_item, user=self.request.user,
                activity_type='STATUS_CHANGED',
                description=f"Status changed from '{old.status.name}' to '{work_item.status.name}'",
                metadata={'old_status': str(old.status_id), 'new_status': str(work_item.status_id)}
            )
            # Auto-set completed_at when moving to done status
            if work_item.status.category == 'done' and not work_item.completed_at:
                work_item.completed_at = timezone.now()
                WorkItem.objects.filter(id=work_item.id).update(completed_at=timezone.now())

            # Phase 2.5: SLA pause/resume on status transition
            if work_item.issue_type == 'TICKET' and work_item.sla_policy:
                waiting_slug = 'waiting_on_customer'
                if work_item.status.slug == waiting_slug:
                    SLAEngine.pause_sla(work_item)
                elif old.status.slug == waiting_slug:
                    SLAEngine.resume_sla(work_item)

        if old.assignee_id != work_item.assignee_id:
            new_name = work_item.assignee.get_full_name() if work_item.assignee else 'Unassigned'
            log_activity(
                work_item=work_item, user=self.request.user,
                activity_type='ASSIGNEE_CHANGED',
                description=f"Assignee changed to {new_name}",
                metadata={
                    'old_assignee': str(old.assignee_id) if old.assignee_id else None,
                    'new_assignee': str(work_item.assignee_id) if work_item.assignee_id else None,
                }
            )

    @action(detail=True, methods=['post'])
    def transition(self, request, pk=None):
        work_item = WorkItem.objects.select_related('status__workflow').get(id=pk)
        status_id = request.data.get('status')

        if not status_id:
            return Response({'error': 'status is required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            new_status = WorkItemStatus.objects.get(id=status_id)
        except WorkItemStatus.DoesNotExist:
            return Response({'error': 'Status not found'}, status=status.HTTP_404_NOT_FOUND)

        old_status = work_item.status

        gate = TransitionRule.objects.filter(
            workflow=work_item.status.workflow,
            from_status=old_status,
            to_status=new_status,
        ).first()
        if gate:
            errors = []
            if gate.require_comment and not request.data.get('comment', '').strip():
                errors.append(gate.error_message or 'A comment is required for this transition')
            if gate.required_field_keys:
                for fk in gate.required_field_keys:
                    val = getattr(work_item, fk, None)
                    if val is None or val == '':
                        errors.append(gate.error_message or f'Field "{fk}" must be set before this transition')
            if gate.required_role:
                user_role = getattr(request.user, 'role', None)
                if str(user_role) != gate.required_role:
                    errors.append(gate.error_message or f'Only users with role "{gate.required_role}" can perform this transition')
            if gate.require_approval:
                has_approval = work_item.approval_requests.filter(
                    status='APPROVED'
                ).exists()
                if not has_approval:
                    errors.append(gate.error_message or 'An approved approval request is required for this transition')
            if errors:
                return Response({'gate_errors': errors}, status=status.HTTP_400_BAD_REQUEST)

        work_item.status = new_status
        work_item.save()

        log_activity(
            work_item=work_item, user=request.user,
            activity_type='STATUS_CHANGED',
            description=f"Status changed from '{old_status.name}' to '{new_status.name}'",
            metadata={'old_status': str(old_status.id), 'new_status': str(new_status.id)}
        )
        return Response({
            'id': str(work_item.id),
            'status': str(work_item.status_id),
            'status_name': new_status.name,
        })

    @action(detail=True, methods=['post'])
    def convert_to_task(self, request, pk=None):
        """Convert a support ticket into a linked dev task."""
        work_item = self.get_object()
        if work_item.issue_type != 'TICKET':
            return Response({'error': 'Only TICKET items can be converted'}, status=400)
        task = WorkItem.objects.create(
            project=work_item.project,
            title=f"[From Ticket] {work_item.title}",
            description=work_item.description,
            issue_type='TASK',
            priority=work_item.priority,
            status=WorkItemStatus.objects.filter(workflow=work_item.project.workflow).first(),
            created_by=request.user,
        )
        WorkItemLink.objects.create(
            source_item=work_item,
            target_item=task,
            relation_type='RELATES_TO',
        )
        log_activity(
            work_item=work_item, user=request.user,
            activity_type='NOTE',
            description=f"Ticket converted to task {task.key}",
        )
        return Response(WorkItemDetailSerializer(task).data, status=201)

    @action(detail=True, methods=['post'])
    def assign(self, request, pk=None):
        work_item = self.get_object()
        user_id = request.data.get('user_id')
        if not user_id:
            return Response({'error': 'user_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)

        old_assignee = work_item.assignee
        work_item.assignee = user
        work_item.save()

        log_activity(
            work_item=work_item, user=request.user,
            activity_type='ASSIGNEE_CHANGED',
            description=f"Assignee changed to {user.get_full_name()}",
            metadata={
                'old_assignee': str(old_assignee.id) if old_assignee else None,
                'new_assignee': str(user.id),
            }
        )
        return Response(WorkItemDetailSerializer(work_item).data)

    @action(detail=True, methods=['get', 'post'])
    def log_time(self, request, pk=None):
        work_item = self.get_object()
        if request.method == 'GET':
            logs = work_item.time_logs.all().select_related('user')
            from project_management.serializers import WorkItemTimeLogSerializer
            return Response(WorkItemTimeLogSerializer(logs, many=True).data)
        hours = request.data.get('hours')
        if not hours:
            return Response({'error': 'hours is required'}, status=status.HTTP_400_BAD_REQUEST)
        log = WorkItemTimeLog.objects.create(
            work_item=work_item,
            user=request.user,
            hours=hours,
            description=request.data.get('description', ''),
            date=request.data.get('date', timezone.now().date()),
        )
        from project_management.serializers import WorkItemTimeLogSerializer
        return Response(WorkItemTimeLogSerializer(log).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['delete'], url_path='log-time/(?P<log_id>[^/.]+)')
    def delete_log_time(self, request, pk=None, log_id=None):
        try:
            log = WorkItemTimeLog.objects.get(id=log_id, work_item_id=pk)
            log.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except WorkItemTimeLog.DoesNotExist:
            return Response({'error': 'Time log not found'}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=['get'])
    def my_items(self, request):
        items = WorkItem.objects.filter(
            assignee=request.user
        ).exclude(status__category='done'
        ).select_related('project', 'status', 'epic', 'sprint'
        ).order_by('due_date', 'priority')

        page = self.paginate_queryset(items)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(items, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['post'], url_path='bulk-status')
    def bulk_update_status(self, request):
        item_ids = request.data.get('item_ids', [])
        status_id = request.data.get('status')
        if not item_ids or not status_id:
            return Response({'error': 'item_ids and status are required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            new_status = WorkItemStatus.objects.get(id=status_id)
        except WorkItemStatus.DoesNotExist:
            return Response({'error': 'Status not found'}, status=status.HTTP_404_NOT_FOUND)
        updated = WorkItem.objects.filter(id__in=item_ids).update(status=new_status)
        return Response({'success': True, 'updated': updated})

    @action(detail=False, methods=['post'], url_path='bulk-reorder')
    def bulk_reorder(self, request):
        items = request.data.get('items', [])
        for item in items:
            WorkItem.objects.filter(id=item['id']).update(order=item['order'])
        return Response({'success': True, 'updated': len(items)})

    @action(detail=False, methods=['post'], url_path='bulk-assign')
    def bulk_assign(self, request):
        item_ids = request.data.get('item_ids', [])
        user_id = request.data.get('user_id')
        if not item_ids or not user_id:
            return Response({'error': 'item_ids and user_id are required'}, status=status.HTTP_400_BAD_REQUEST)
        updated = WorkItem.objects.filter(id__in=item_ids).update(assignee_id=user_id)
        return Response({'success': True, 'updated': updated})

    @action(detail=False, methods=['post'], url_path='bulk-edit')
    def bulk_edit(self, request):
        item_ids = request.data.get('item_ids', [])
        changes = request.data.get('changes', {})
        if not item_ids:
            return Response({'error': 'item_ids is required'}, status=status.HTTP_400_BAD_REQUEST)
        allowed = {'sprint_id', 'priority', 'issue_type', 'milestone_id', 'epic_id', 'story_points', 'estimated_hours'}
        update_kwargs = {}
        for key, val in changes.items():
            mapped = {'sprint_id': 'sprint_id', 'milestone_id': 'milestone_id', 'epic_id': 'epic_id'}
            field = mapped.get(key, key)
            if field in allowed:
                update_kwargs[field] = val if val != '' else None
        if not update_kwargs:
            return Response({'error': 'No valid changes provided'}, status=status.HTTP_400_BAD_REQUEST)
        updated = WorkItem.objects.filter(id__in=item_ids).update(**update_kwargs)
        return Response({'success': True, 'updated': updated})

    @action(detail=False, methods=['post'], url_path='bulk-delete')
    def bulk_delete(self, request):
        item_ids = request.data.get('item_ids', [])
        if not item_ids:
            return Response({'error': 'item_ids is required'}, status=status.HTTP_400_BAD_REQUEST)
        deleted, _ = WorkItem.objects.filter(id__in=item_ids).delete()
        return Response({'success': True, 'deleted': deleted})

    @action(detail=True, methods=['post'], url_path='add-to-sprint')
    def add_to_sprint(self, request, pk=None):
        """Assign a work item to a sprint."""
        work_item = self.get_object()
        sprint_id = request.data.get('sprint_id')
        if not sprint_id:
            return Response({'error': 'sprint_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            sprint = Sprint.objects.get(id=sprint_id, project=work_item.project)
        except Sprint.DoesNotExist:
            return Response({'error': 'Sprint not found'}, status=status.HTTP_404_NOT_FOUND)

        work_item.sprint = sprint
        work_item.save()
        return Response(WorkItemDetailSerializer(work_item).data)

    @action(detail=True, methods=['post'], url_path='remove-from-sprint')
    def remove_from_sprint(self, request, pk=None):
        """Remove a work item from its sprint (back to backlog)."""
        work_item = self.get_object()
        work_item.sprint = None
        work_item.save()
        return Response(WorkItemDetailSerializer(work_item).data)

    @action(detail=True, methods=['post'], url_path='update-points')
    def update_points(self, request, pk=None):
        """Update story points estimation."""
        work_item = self.get_object()
        points = request.data.get('story_points')
        if points is None:
            return Response({'error': 'story_points is required'}, status=status.HTTP_400_BAD_REQUEST)
        work_item.story_points = points
        work_item.save()
        return Response(WorkItemDetailSerializer(work_item).data)

    # ═══════════════════════════════════════════════════════════════════
    # Phase 1: Release/Version tagging
    # ═══════════════════════════════════════════════════════════════════

    @action(detail=True, methods=['post'], url_path='tag-version')
    def tag_version(self, request, pk=None):
        """Tag a work item to a release/version."""
        work_item = self.get_object()
        version_id = request.data.get('version_id')
        if not version_id:
            return Response({'error': 'version_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            release = Release.objects.get(id=version_id, project=work_item.project)
        except Release.DoesNotExist:
            return Response({'error': 'Release not found'}, status=status.HTTP_404_NOT_FOUND)
        work_item.version = release
        work_item.save()
        return Response(WorkItemDetailSerializer(work_item).data)

    @action(detail=True, methods=['post'], url_path='remove-version')
    def remove_version(self, request, pk=None):
        """Remove release/version tag from a work item."""
        work_item = self.get_object()
        work_item.version = None
        work_item.save()
        return Response(WorkItemDetailSerializer(work_item).data)

    # ═══════════════════════════════════════════════════════════════════
    # Phase 2.5: Ticketing Mode — SLA Actions
    # ═══════════════════════════════════════════════════════════════════

    @action(detail=True, methods=['get'])
    def check_sla(self, request, pk=None):
        """Check SLA status for a ticket."""
        work_item = self.get_object()
        result = SLAEngine.check_sla(work_item)
        return Response(result)

    @action(detail=True, methods=['post'])
    def mark_responded(self, request, pk=None):
        """Mark that an agent has responded to a ticket (first response tracking)."""
        work_item = self.get_object()
        SLAEngine.mark_first_response(work_item, request.user)
        return Response(WorkItemDetailSerializer(work_item).data)

    @action(detail=True, methods=['post'])
    def start_sla(self, request, pk=None):
        """Manually start/restart SLA timer for a ticket."""
        work_item = self.get_object()
        policy_id = request.data.get('sla_policy_id')
        if policy_id:
            try:
                work_item.sla_policy = SLAPolicy.objects.get(id=policy_id)
                work_item.save(update_fields=['sla_policy'])
            except SLAPolicy.DoesNotExist:
                return Response({'error': 'SLA Policy not found'}, status=status.HTTP_404_NOT_FOUND)
        SLAEngine.start_sla(work_item)
        return Response(WorkItemDetailSerializer(work_item).data)

    @action(detail=True, methods=['post'])
    def pause_sla(self, request, pk=None):
        """Pause SLA timer (e.g., waiting on customer)."""
        work_item = self.get_object()
        SLAEngine.pause_sla(work_item)
        return Response(WorkItemDetailSerializer(work_item).data)

    @action(detail=True, methods=['post'])
    def resume_sla(self, request, pk=None):
        """Resume SLA timer after pause."""
        work_item = self.get_object()
        SLAEngine.resume_sla(work_item)
        return Response(WorkItemDetailSerializer(work_item).data)

    @action(detail=True, methods=['post'])
    def submit_csat(self, request, pk=None):
        """Submit a CSAT survey response for a resolved ticket."""
        work_item = self.get_object()
        rating = request.data.get('rating')
        if not rating or not (1 <= int(rating) <= 5):
            return Response({'error': 'rating must be 1-5'}, status=status.HTTP_400_BAD_REQUEST)

        csat = CSATResponse.objects.create(
            work_item=work_item,
            rating=int(rating),
            comment=request.data.get('comment', ''),
            responded_by=request.data.get('responded_by', ''),
        )

        log_activity(
            work_item=work_item, user=request.user,
            activity_type='CSAT_SUBMITTED',
            description=f"CSAT rating: {rating}/5",
            metadata={'rating': rating, 'csat_id': str(csat.id)}
        )

        return Response(CSATResponseSerializer(csat).data)

    # ── Phase 4: Import / Export ──────────────────────────────────────────

    @action(detail=False, methods=['post'])
    def import_csv(self, request):
        """Import work items from a CSV file. Expects multipart/form-data with a 'file' field.
        Required CSV columns: title, issue_type (or defaults to TASK).
        Optional: description, priority, due_date, story_points, labels.
        """
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

        project_id = request.data.get('project_id')
        if not project_id:
            return Response({'error': 'project_id is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            project = Project.objects.get(id=project_id)
        except Project.DoesNotExist:
            return Response({'error': 'Project not found'}, status=status.HTTP_404_NOT_FOUND)

        import csv, io
        decoded = file.read().decode('utf-8')
        reader = csv.DictReader(io.StringIO(decoded))

        created = []
        errors = []

        status_obj = WorkItemStatus.objects.filter(project=project).first()
        if not status_obj:
            return Response({'error': 'No status configured for this project'}, status=status.HTTP_400_BAD_REQUEST)

        for i, row in enumerate(reader, start=1):
            title = row.get('title', '').strip()
            if not title:
                errors.append({'row': i, 'error': 'title is required'})
                continue
            try:
                item = WorkItem.objects.create(
                    project=project,
                    title=title,
                    issue_type=row.get('issue_type', 'TASK').strip().upper(),
                    description=row.get('description', '').strip(),
                    priority=row.get('priority', 'MEDIUM').strip().upper(),
                    status=status_obj,
                    story_points=int(row['story_points']) if row.get('story_points') else None,
                    labels=[l.strip() for l in row.get('labels', '').split(',') if l.strip()] if row.get('labels') else [],
                    reporter=request.user,
                )
                created.append({
                    'key': item.key,
                    'title': item.title,
                    'id': str(item.id),
                })
            except Exception as e:
                errors.append({'row': i, 'error': str(e)})

        return Response({
            'created_count': len(created),
            'error_count': len(errors),
            'created': created,
            'errors': errors,
        })

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Export all work items for a project as an Excel (.xlsx) file."""
        project_id = request.query_params.get('project_id')
        if not project_id:
            return Response({'error': 'project_id query param is required'}, status=status.HTTP_400_BAD_REQUEST)

        items = self.get_queryset().filter(project_id=project_id).select_related(
            'assignee', 'status', 'sprint', 'sla_policy'
        )

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse

        wb = Workbook()
        ws = wb.active
        ws.title = 'Work Items'

        headers = ['Key', 'Title', 'Issue Type', 'Status', 'Priority', 'Assignee',
                   'Story Points', 'Sprint', 'Due Date', 'Labels', 'SLA Status']
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2D3748', end_color='2D3748', fill_type='solid')

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for row_idx, item in enumerate(items, start=2):
            ws.cell(row=row_idx, column=1, value=item.key)
            ws.cell(row=row_idx, column=2, value=item.title)
            ws.cell(row=row_idx, column=3, value=item.issue_type)
            ws.cell(row=row_idx, column=4, value=item.status.name if item.status else '')
            ws.cell(row=row_idx, column=5, value=item.priority)
            ws.cell(row=row_idx, column=6, value=item.assignee.get_full_name() if item.assignee else '')
            ws.cell(row=row_idx, column=7, value=item.story_points)
            ws.cell(row=row_idx, column=8, value=item.sprint.name if item.sprint else '')
            ws.cell(row=row_idx, column=9, value=item.due_date.strftime('%Y-%m-%d') if item.due_date else '')
            ws.cell(row=row_idx, column=10, value=','.join(item.labels) if item.labels else '')
            ws.cell(row=row_idx, column=11, value=item.sla_status or '')

        from django.utils import timezone
        filename = f"work_items_{project_id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


# ============================================================================
# WORK ITEM LINK VIEWSET
# ============================================================================

class WorkItemLinkViewSet(viewsets.ModelViewSet):
    queryset = WorkItemLink.objects.all().select_related('source_item', 'target_item')
    serializer_class = WorkItemLinkSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['source_item', 'target_item', 'relation_type']


# ============================================================================
# WORK ITEM COMMENT VIEWSET
# ============================================================================

class WorkItemCommentViewSet(viewsets.ModelViewSet):
    queryset = WorkItemComment.objects.all().select_related('author')
    serializer_class = WorkItemCommentSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['work_item', 'author', 'parent']

    def perform_create(self, serializer):
        comment = serializer.save(author=self.request.user)
        log_activity(
            work_item=comment.work_item, user=self.request.user,
            activity_type='COMMENT_ADDED',
            description=f"Comment added by {self.request.user.get_full_name()}",
            metadata={'comment_id': str(comment.id)}
        )


# ============================================================================
# WORK ITEM ATTACHMENT VIEWSET
# ============================================================================

class WorkItemAttachmentViewSet(viewsets.ModelViewSet):
    queryset = WorkItemAttachment.objects.all().select_related('uploaded_by', 'work_item')
    serializer_class = WorkItemAttachmentSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['work_item', 'uploaded_by']

    def perform_create(self, serializer):
        attachment = serializer.save(uploaded_by=self.request.user)
        log_activity(
            work_item=attachment.work_item, user=self.request.user,
            activity_type='ATTACHMENT_ADDED',
            description=f"Attachment '{attachment.file_name}' added",
            metadata={'attachment_id': str(attachment.id), 'file_name': attachment.file_name}
        )


# ============================================================================
# ACTIVITY LOG VIEWSET
# ============================================================================

class WorkItemActivityLogViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = WorkItemActivityLog.objects.all().select_related('user')
    serializer_class = WorkItemActivityLogSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['work_item', 'user', 'activity_type']

    def get_queryset(self):
        user = self.request.user
        qs = super().get_queryset()
        if user.is_authenticated and user.role == 'Staff':
            qs = qs.filter(work_item__assignee=user) | qs.filter(user=user)
        return qs.distinct()


# ============================================================================
# SLA POLICY VIEWSET — Phase 2.5: Ticketing Mode
# ============================================================================

class SLAPolicyViewSet(viewsets.ModelViewSet):
    queryset = SLAPolicy.objects.all().select_related('project')
    serializer_class = SLAPolicySerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['project', 'priority', 'is_default']
    search_fields = ['name']

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [CanManageWorkflows()]
        return [permissions.IsAuthenticated()]


# ============================================================================
# CSAT RESPONSE VIEWSET — Phase 2.5: Ticketing Mode
# ============================================================================

class CSATResponseViewSet(viewsets.ModelViewSet):
    queryset = CSATResponse.objects.all().select_related('work_item')
    serializer_class = CSATResponseSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['work_item', 'rating']


# ============================================================================
# MILESTONE VIEWSET — Phase 3: Cross-Cutting Features
# ============================================================================

class MilestoneViewSet(viewsets.ModelViewSet):
    queryset = Milestone.objects.all().select_related(
        'project', 'sprint', 'work_item'
    )
    serializer_class = MilestoneSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['project', 'sprint', 'work_item', 'status', 'milestone_type']
    search_fields = ['name', 'description']

    @action(detail=True, methods=['post'])
    def achieve(self, request, pk=None):
        milestone = self.get_object()
        milestone.status = 'ACHIEVED'
        milestone.completed_date = timezone.now()
        milestone.save()

        log_activity(
            work_item=milestone.work_item,
            user=request.user,
            activity_type='MILESTONE_ACHIEVED',
            description=f"Milestone '{milestone.name}' achieved",
            metadata={
                'milestone_id': str(milestone.id),
                'milestone_type': milestone.milestone_type,
            }
        )
        return Response(MilestoneSerializer(milestone).data)

    @action(detail=True, methods=['post'])
    def miss(self, request, pk=None):
        milestone = self.get_object()
        milestone.status = 'MISSED'
        milestone.save()

        log_activity(
            work_item=milestone.work_item,
            user=request.user,
            activity_type='MILESTONE_MISSED',
            description=f"Milestone '{milestone.name}' missed",
            metadata={'milestone_id': str(milestone.id)}
        )
        return Response(MilestoneSerializer(milestone).data)

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy', 'achieve', 'miss']:
            return [IsProjectAdminOrEditor()]
        return [permissions.IsAuthenticated()]


# ============================================================================
# OKR / GOAL TREE VIEWSETS (Phase 3.2)
# ============================================================================

class ObjectiveViewSet(viewsets.ModelViewSet):
    queryset = Objective.objects.all().select_related('owner', 'parent_objective', 'project')
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['workspace', 'project', 'status', 'alignment', 'owner']
    search_fields = ['title', 'description']

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return ObjectiveDetailSerializer
        return ObjectiveListSerializer

    @action(detail=True, methods=['post'])
    def update_progress(self, request, pk=None):
        """Recalculate objective progress from all key results."""
        objective = self.get_object()
        krs = objective.key_results.all()
        if not krs:
            return Response({'avg_progress': 0, 'message': 'No key results'})
        avg = round(sum(kr.progress_pct for kr in krs) / len(krs), 1)
        return Response({'avg_progress': avg, 'key_results_count': len(krs)})

    @action(detail=True, methods=['post'])
    def add_key_result(self, request, pk=None):
        objective = self.get_object()
        serializer = KeyResultSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(objective=objective)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class KeyResultViewSet(viewsets.ModelViewSet):
    queryset = KeyResult.objects.all().select_related('objective', 'metric_work_item')
    serializer_class = KeyResultSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['objective']


# ============================================================================
# AUTOMATION RULE VIEWSET (Phase 3.4)
# ============================================================================

class AutomationRuleViewSet(viewsets.ModelViewSet):
    queryset = AutomationRule.objects.all().select_related('project')
    serializer_class = AutomationRuleSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['project', 'trigger_event', 'action_type', 'is_enabled', 'issue_type_filter']
    search_fields = ['name', 'description']

    @action(detail=True, methods=['post'])
    def toggle(self, request, pk=None):
        rule = self.get_object()
        rule.is_enabled = not rule.is_enabled
        rule.save(update_fields=['is_enabled'])
        return Response({'is_enabled': rule.is_enabled})

    @action(detail=True, methods=['post'])
    def test(self, request, pk=None):
        """Test a rule against a specific work item."""
        rule = self.get_object()
        work_item_id = request.data.get('work_item_id')
        if not work_item_id:
            return Response({'error': 'work_item_id required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            from project_management.models import WorkItem
            item = WorkItem.objects.get(id=work_item_id)
            matches = rule.evaluate_conditions(item)
            return Response({
                'rule_name': rule.name,
                'work_item': item.key,
                'conditions_match': matches,
                'trigger_event': rule.trigger_event,
                'action': rule.action_type,
                'action_config': rule.action_config,
            })
        except WorkItem.DoesNotExist:
            return Response({'error': 'WorkItem not found'}, status=status.HTTP_404_NOT_FOUND)


# ============================================================================
# NOTIFICATION VIEWSETS (Phase 3.5)
# ============================================================================

class NotificationViewSet(viewsets.ModelViewSet):
    queryset = Notification.objects.all().select_related('recipient', 'work_item', 'project')
    serializer_class = NotificationSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['recipient', 'notification_type', 'is_read', 'project']

    def get_queryset(self):
        qs = super().get_queryset()
        if not self.request.user.is_superuser:
            qs = qs.filter(recipient=self.request.user)
        return qs

    @action(detail=True, methods=['post'])
    def mark_read(self, request, pk=None):
        notification = self.get_object()
        notification.mark_read()
        return Response({'is_read': True, 'read_at': notification.read_at})

    @action(detail=False, methods=['post'])
    def mark_all_read(self, request):
        count = Notification.objects.filter(
            recipient=request.user, is_read=False
        ).update(is_read=True, read_at=timezone.now())
        return Response({'marked_read': count})

    @action(detail=False, methods=['get'])
    def unread_count(self, request):
        count = Notification.objects.filter(
            recipient=request.user, is_read=False
        ).count()
        return Response({'count': count})

    @action(detail=False, methods=['get'])
    def smart_list(self, request):
        """Grouped/digest notifications for smart inbox."""
        from project_management.services.notification_engine import get_smart_notifications
        project = request.query_params.get('project')
        days = int(request.query_params.get('days', 7))
        data = get_smart_notifications(user=request.user, project_id=project, days=days)
        return Response(data)


class NotificationPreferenceViewSet(viewsets.ModelViewSet):
    queryset = NotificationPreference.objects.all().select_related('user', 'project')
    serializer_class = NotificationPreferenceSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['project', 'notification_type', 'channel', 'enabled']

    def get_queryset(self):
        qs = super().get_queryset()
        return qs.filter(user=self.request.user)


class WebhookConfigViewSet(viewsets.ModelViewSet):
    queryset = WebhookConfig.objects.all().select_related('project')
    serializer_class = WebhookConfigSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['project', 'provider', 'is_enabled']

    @action(detail=True, methods=['post'])
    def test(self, request, pk=None):
        """Send a test message to the webhook."""
        config = self.get_object()
        from project_management.services.notification_engine import NotificationEngine
        success = NotificationEngine.send_slack(
            config.webhook_url,
            f"🔔 Test notification from HertexFlow — {config.name or 'Webhook'} is working!",
        )
        return Response({
            'success': success,
            'message': 'Test sent successfully' if success else 'Failed to send test',
        })


# ============================================================================
# CUSTOM FIELD VIEWSET (Phase 3.6)
# ============================================================================

class CustomFieldDefinitionViewSet(viewsets.ModelViewSet):
    queryset = CustomFieldDefinition.objects.all().select_related('project')
    serializer_class = CustomFieldDefinitionSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['project', 'issue_type_filter', 'field_type', 'is_active']
    search_fields = ['label', 'field_key']


# ============================================================================
# FIELD VISIBILITY VIEWSET (Phase 4.3)
# ============================================================================

class ProjectFieldVisibilityViewSet(viewsets.ModelViewSet):
    queryset = ProjectFieldVisibility.objects.all().select_related('project')
    serializer_class = ProjectFieldVisibilitySerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['project', 'role']


# ============================================================================
# DASHBOARD VIEWSET
# ============================================================================

class DashboardViewSet(viewsets.ViewSet):
    permission_classes = [permissions.IsAuthenticated]

    @action(detail=False, methods=['get'])
    def my_work(self, request):
        my_items = WorkItem.objects.filter(assignee=request.user)
        overdue = my_items.filter(
            due_date__lt=timezone.now(),
            status__category__in=['todo', 'in_progress']
        )
        due_today = my_items.filter(
            due_date__date=timezone.now().date()
        ).exclude(status__category='done')
        active_sprint_items = my_items.filter(sprint__status='ACTIVE')

        return Response({
            'total_assigned': my_items.count(),
            'done': my_items.filter(status__category='done').count(),
            'in_progress': my_items.filter(status__category='in_progress').count(),
            'todo': my_items.filter(status__category='todo').count(),
            'overdue': overdue.count(),
            'due_today': due_today.count(),
            'in_active_sprint': active_sprint_items.count(),
            'projects': list(my_items.values('project__name', 'project_id').annotate(
                count=Count('id')
            ).order_by('-count')),
        })

    @action(detail=False, methods=['get'])
    def workspace_overview(self, request):
        workspaces = Workspace.objects.annotate(
            project_count=Count('projects', distinct=True),
        )
        data = []
        for ws in workspaces:
            items = WorkItem.objects.filter(project__workspace=ws)
            active_sprints = Sprint.objects.filter(
                project__workspace=ws, status='ACTIVE'
            ).count()
            data.append({
                'id': str(ws.id),
                'name': ws.name,
                'project_count': ws.project_count,
                'total_items': items.count(),
                'open_items': items.exclude(status__category='done').count(),
                'overdue': items.filter(
                    due_date__lt=timezone.now(),
                    status__category__in=['todo', 'in_progress']
                ).count(),
                'active_sprints': active_sprints,
            })
        return Response(data)

    # ═══════════════════════════════════════════════════════════════════
    # Phase 2: Sales Dashboard Endpoints
    # ═══════════════════════════════════════════════════════════════════

    @action(detail=False, methods=['get'])
    def sales_attainment(self, request):
        """Target attainment data from sales_task_manager for active cycles."""
        if not SALES_TASK_MANAGER_AVAILABLE:
            return Response({'error': 'sales_task_manager not installed'}, status=501)

        cycle_id = request.query_params.get('cycle')
        workspace_id = request.query_params.get('workspace')

        cycles = TargetCycle.objects.filter(status='ACTIVE')
        if cycle_id:
            cycles = cycles.filter(id=cycle_id)

        data = []
        for cycle in cycles:
            targets = SalesTarget.objects.filter(cycle=cycle).select_related(
                'assigned_user', 'assigned_department'
            )

            target_data = []
            for t in targets:
                target_data.append({
                    'id': str(t.id),
                    'assignee_type': t.assignee_type,
                    'assigned_user': UserBriefSerializer(t.assigned_user).data if t.assigned_user else None,
                    'assigned_department': t.assigned_department.name if t.assigned_department else None,
                    'target_amount': float(t.target_amount),
                    'achieved_amount': float(t.achieved_amount),
                    'weighted_progress_pct': float(t.weighted_progress_pct),
                    'attainment_pct': round(
                        (float(t.achieved_amount) / float(t.target_amount) * 100)
                        if t.target_amount else 0, 1
                    ),
                    'status': t.status,
                })

            total_target = sum(t['target_amount'] for t in target_data)
            total_achieved = sum(t['achieved_amount'] for t in target_data)

            data.append({
                'cycle': {
                    'id': str(cycle.id),
                    'name': cycle.name,
                    'code': cycle.code,
                    'cycle_type': cycle.cycle_type,
                    'start_date': cycle.start_date.isoformat(),
                    'end_date': cycle.end_date.isoformat(),
                    'status': cycle.status,
                },
                'total_target': total_target,
                'total_achieved': total_achieved,
                'overall_attainment_pct': round(
                    (total_achieved / total_target * 100) if total_target else 0, 1
                ),
                'targets': target_data,
                'target_count': len(target_data),
            })

        return Response(data)

    @action(detail=False, methods=['get'])
    def sales_leaderboard(self, request):
        """User ranking by achieved revenue from sales targets."""
        if not SALES_TASK_MANAGER_AVAILABLE:
            return Response({'error': 'sales_task_manager not installed'}, status=501)

        cycle_id = request.query_params.get('cycle')
        limit = int(request.query_params.get('limit', 10))

        targets = SalesTarget.objects.filter(
            assignee_type='USER',
            assigned_user__isnull=False,
        ).select_related('assigned_user', 'cycle')

        if cycle_id:
            targets = targets.filter(cycle_id=cycle_id)
        else:
            targets = targets.filter(cycle__status='ACTIVE')

        from django.db.models import Count, Sum, DecimalField
        from django.db.models.functions import Coalesce

        rankings = targets.values(
            'assigned_user', 'assigned_user__first_name',
            'assigned_user__last_name', 'assigned_user__email',
            'assigned_user__role',
        ).annotate(
            total_target=Coalesce(Sum('target_amount'), 0, output_field=DecimalField()),
            total_achieved=Coalesce(Sum('achieved_amount'), 0, output_field=DecimalField()),
            target_count=Count('id'),
        ).order_by('-total_achieved')[:limit]

        leaderboard = []
        for rank, r in enumerate(rankings, 1):
            target_val = float(r['total_target'])
            achieved_val = float(r['total_achieved'])
            leaderboard.append({
                'rank': rank,
                'user_id': str(r['assigned_user']),
                'name': f"{r['assigned_user__first_name']} {r['assigned_user__last_name']}".strip() or r['assigned_user__email'],
                'role': r['assigned_user__role'],
                'total_target': target_val,
                'total_achieved': achieved_val,
                'attainment_pct': round(
                    (achieved_val / target_val * 100) if target_val else 0, 1
                ),
                'target_count': r['target_count'],
            })

        return Response(leaderboard)

    @action(detail=False, methods=['get'])
    def sales_forecast(self, request):
        """Weighted pipeline forecast from DEAL work items."""
        project_id = request.query_params.get('project')
        workspace_id = request.query_params.get('workspace')

        deals = WorkItem.objects.filter(issue_type='DEAL').exclude(
            status__category__in=['done', 'cancelled']
        ).select_related('project', 'status', 'assignee')

        if project_id:
            deals = deals.filter(project_id=project_id)
        elif workspace_id:
            deals = deals.filter(project__workspace_id=workspace_id)

        stage_probability = {
            'lead': 10,
            'qualified': 25,
            'proposal_sent': 50,
            'negotiation': 75,
            'won': 100,
            'lost': 0,
        }

        stage_breakdown = {}
        total_weighted = 0
        total_pipeline = 0
        deal_list = []

        for deal in deals:
            cf = deal.custom_fields or {}
            deal_value = float(cf.get('deal_value', 0))
            prob_slug = deal.status.slug.lower()
            probability = stage_probability.get(prob_slug, float(cf.get('probability', 10)))
            weighted = round(deal_value * probability / 100, 2)
            total_pipeline += deal_value
            total_weighted += weighted

            stage_name = deal.status.name
            if stage_name not in stage_breakdown:
                stage_breakdown[stage_name] = {
                    'deal_count': 0,
                    'total_value': 0,
                    'weighted_value': 0,
                    'probability': probability,
                }
            stage_breakdown[stage_name]['deal_count'] += 1
            stage_breakdown[stage_name]['total_value'] += deal_value
            stage_breakdown[stage_name]['weighted_value'] += weighted

            deal_list.append({
                'id': str(deal.id),
                'key': deal.key,
                'title': deal.title,
                'deal_value': deal_value,
                'probability': probability,
                'weighted_value': weighted,
                'stage': stage_name,
                'assignee': UserBriefSerializer(deal.assignee).data if deal.assignee else None,
                'expected_close_date': cf.get('expected_close_date'),
                'project_key': deal.project.key,
            })

        return Response({
            'total_pipeline_value': total_pipeline,
            'total_weighted_forecast': round(total_weighted, 2),
            'deal_count': len(deal_list),
            'stage_breakdown': stage_breakdown,
            'deals': sorted(deal_list, key=lambda x: x['weighted_value'], reverse=True),
        })

    @action(detail=False, methods=['get'])
    def sales_pipeline_summary(self, request):
        """Summary of all DEAL pipelines across projects in a workspace."""
        workspace_id = request.query_params.get('workspace')

        projects = Project.objects.all()
        if workspace_id:
            projects = projects.filter(workspace_id=workspace_id)

        pipeline_data = []
        grand_total_value = 0
        grand_total_weighted = 0
        grand_total_deals = 0

        for project in projects:
            deals = WorkItem.objects.filter(
                project=project, issue_type='DEAL'
            ).exclude(status__category__in=['done', 'cancelled'])

            if not deals.exists():
                continue

            total_value = 0
            total_weighted = 0
            for d in deals:
                cf = d.custom_fields or {}
                dv = float(cf.get('deal_value', 0))
                prob = float(cf.get('probability', 10))
                total_value += dv
                total_weighted += round(dv * prob / 100, 2)

            pipeline_data.append({
                'project_id': str(project.id),
                'project_name': project.name,
                'project_key': project.key,
                'deal_count': deals.count(),
                'total_value': total_value,
                'weighted_forecast': round(total_weighted, 2),
            })
            grand_total_value += total_value
            grand_total_weighted += total_weighted
            grand_total_deals += deals.count()

        return Response({
            'pipelines': pipeline_data,
            'summary': {
                'total_projects': len(pipeline_data),
                'total_deals': grand_total_deals,
                'total_pipeline_value': grand_total_value,
                'total_weighted_forecast': round(grand_total_weighted, 2),
            },
        })

    # ═══════════════════════════════════════════════════════════════════
    # Phase 3: Executive Dashboard
    # ═══════════════════════════════════════════════════════════════════

    @action(detail=False, methods=['get'])
    def executive(self, request):
        """Cross-department executive dashboard — dev velocity + sales attainment + ticket SLAs."""
        workspace_id = request.query_params.get('workspace')

        projects = Project.objects.all()
        if workspace_id:
            projects = projects.filter(workspace_id=workspace_id)

        # ── 1. Dev Velocity ─────────────────────────────────────────────
        completed_sprints = Sprint.objects.filter(
            project__in=projects, status='COMPLETED'
        ).select_related('project').order_by('-end_date')[:10]

        velocity_data = []
        total_velocity = 0
        for sprint in completed_sprints:
            items = WorkItem.objects.filter(sprint=sprint)
            committed = items.aggregate(total=Sum('story_points'))['total'] or 0
            completed = items.filter(
                status__category='done'
            ).aggregate(total=Sum('story_points'))['total'] or 0
            velocity_data.append({
                'sprint_name': sprint.name,
                'project_key': sprint.project.key,
                'committed': float(committed),
                'completed': float(completed),
                'completion_pct': round(
                    float(completed) / float(committed) * 100 if committed else 0, 1
                ),
            })
            total_velocity += float(completed)

        avg_velocity = round(
            total_velocity / len(velocity_data), 1
        ) if velocity_data else 0

        active_sprints_count = Sprint.objects.filter(
            project__in=projects, status='ACTIVE'
        ).count()

        # ── 2. Sales Attainment ─────────────────────────────────────────
        sales_summary = {
            'total_target': 0, 'total_achieved': 0,
            'overall_attainment': 0, 'active_cycles': 0,
        }
        if SALES_TASK_MANAGER_AVAILABLE:
            active_cycles = TargetCycle.objects.filter(status='ACTIVE')
            all_targets = SalesTarget.objects.filter(cycle__in=active_cycles)
            attain_data = all_targets.aggregate(
                total_target=Sum('target_amount'),
                total_achieved=Sum('achieved_amount'),
            )
            t_target = float(attain_data['total_target'] or 0)
            t_achieved = float(attain_data['total_achieved'] or 0)
            sales_summary = {
                'total_target': t_target,
                'total_achieved': t_achieved,
                'overall_attainment': round(
                    (t_achieved / t_target * 100) if t_target else 0, 1
                ),
                'active_cycles': active_cycles.count(),
            }

        # ── 3. Ticket SLA Stats ─────────────────────────────────────────
        tickets = WorkItem.objects.filter(
            project__in=projects, issue_type='TICKET'
        ).exclude(status__category='done')

        sla_breached = tickets.filter(sla_status='BREACHED').count()
        sla_warning = tickets.filter(sla_status='WARNING').count()
        sla_within = tickets.filter(sla_status='WITHIN_SLA').count()
        total_tickets = tickets.count()

        # ── 4. Overall Stats ────────────────────────────────────────────
        all_items = WorkItem.objects.filter(project__in=projects)
        total_items = all_items.count()
        done_items = all_items.filter(status__category='done').count()
        in_progress = all_items.filter(status__category='in_progress').count()
        todo_items = all_items.filter(status__category='todo').count()

        overdue = all_items.filter(
            due_date__lt=timezone.now(),
            status__category__in=['todo', 'in_progress']
        ).count()

        unassigned = all_items.filter(assignee__isnull=True).count()

        # ── 5. Milestones ───────────────────────────────────────────────
        upcoming_milestones = Milestone.objects.filter(
            project__in=projects, status='PENDING',
            target_date__gte=timezone.now(),
            target_date__lte=timezone.now() + timezone.timedelta(days=30),
        ).count()

        achieved_milestones = Milestone.objects.filter(
            project__in=projects, status='ACHIEVED'
        ).count()

        return Response({
            'workspace_id': workspace_id,
            'projects_count': projects.count(),
            'dev': {
                'avg_velocity': avg_velocity,
                'velocity_trend': velocity_data[:5],
                'active_sprints': active_sprints_count,
                'sprints_analyzed': len(velocity_data),
            },
            'sales': sales_summary,
            'tickets': {
                'total_open': total_tickets,
                'sla_breached': sla_breached,
                'sla_warning': sla_warning,
                'sla_within': sla_within,
                'sla_health_pct': round(
                    (sla_within / total_tickets * 100) if total_tickets else 100, 1
                ),
            },
            'overview': {
                'total_items': total_items,
                'done': done_items,
                'in_progress': in_progress,
                'todo': todo_items,
                'overdue': overdue,
                'unassigned': unassigned,
                'completion_pct': round(
                    (done_items / total_items * 100) if total_items else 0, 1
                ),
            },
            'milestones': {
                'upcoming_30_days': upcoming_milestones,
                'achieved_total': achieved_milestones,
            },
        })

    # ═══════════════════════════════════════════════════════════════════
    # Phase 4: Organization-wide Business Owner Dashboard
    # ═══════════════════════════════════════════════════════════════════

    @action(detail=False, methods=['get'])
    def owner(self, request):
        """Single cross-module rollup for the business owner — the whole
        organization at a glance: people, CRM, finance, sales, projects,
        inventory, calendar and a merged recent-activity feed."""
        from django.db.models.functions import TruncMonth

        User = get_user_model()
        now = timezone.now()

        # ── Combined single-round-trip rollup ───────────────────────────
        # The owner dashboard aggregates many counters across modules.
        # Neon's pooled connection adds ~0.5s of latency per query, so
        # issuing ~50 separate count/aggregate queries here made the whole
        # endpoint take >20s and time out (HTTP 500). All scalar rollups
        # below are fetched in a single query instead.
        from django.db import connection as db_connection

        agg_parts = []   # list of (key, sql)
        agg_params = []  # positional params, in the same order as parts

        def _add(key, sql, params=()):
            agg_parts.append((key, f"SELECT {sql}"))
            agg_params.extend(params)

        user_tbl = User._meta.db_table
        _add('user_total', f"COUNT(*) FROM {user_tbl}")
        _add('user_active', f"COUNT(*) FROM {user_tbl} WHERE is_active")

        try:
            from authentication.models import Department
            _add('dept_count', f"COUNT(*) FROM {Department._meta.db_table}")
        except ImportError:
            pass

        try:
            from hr.models import (
                Employee, LeaveApplication, Attendance, Payroll,
            )
            emp_tbl = Employee._meta.db_table
            _add('emp_total', f"COUNT(*) FROM {emp_tbl}")
            _add('emp_active', f"COUNT(*) FROM {emp_tbl} WHERE is_active")
            _add('emp_onboarding', f"COUNT(*) FROM {emp_tbl} WHERE status = 'ONBOARDING'")
            _add('emp_notice', f"COUNT(*) FROM {emp_tbl} WHERE status = 'NOTICE_PERIOD'")
            _add('leave_pending', f"COUNT(*) FROM {LeaveApplication._meta.db_table} WHERE approval_status = 'PENDING'")
            today = now.date()
            _add('attendance_today', f"COUNT(*) FROM {Attendance._meta.db_table} WHERE date = %s AND status = 'PRESENT'", [today])
            _add('payroll_total', f"COALESCE(SUM(gross_salary), 0) FROM {Payroll._meta.db_table} WHERE status <> 'DRAFT'")
        except ImportError:
            pass

        try:
            from contacts.models import Contact
            c_tbl = Contact._meta.db_table
            _add('contact_total', f"COUNT(*) FROM {c_tbl}")
            _add('contact_new', f"COUNT(*) FROM {c_tbl} WHERE created_at >= %s", [now - timezone.timedelta(days=30)])
        except ImportError:
            pass

        try:
            from crm.models import CRM, Stage
            crm_tbl = CRM._meta.db_table
            stage_tbl = Stage._meta.db_table
            _add('crm_deals', f"COUNT(*) FROM {crm_tbl}")
            _add('crm_pipeline', f"COALESCE(SUM(value), 0) FROM {crm_tbl}")
            _add('crm_won', f"COUNT(*) FROM {crm_tbl} c JOIN {stage_tbl} s ON c.stage_id = s.id WHERE s.slug IN ('won', 'closed_won')")
            _add('crm_won_value', f"COALESCE(SUM(c.value), 0) FROM {crm_tbl} c JOIN {stage_tbl} s ON c.stage_id = s.id WHERE s.slug IN ('won', 'closed_won')")
        except ImportError:
            pass

        try:
            from invoices.models import Invoice
            inv_tbl = Invoice._meta.db_table
            _add('inv_total', f"COUNT(*) FROM {inv_tbl}")
            _add('inv_invoiced', f"COALESCE(SUM(grand_total), 0) FROM {inv_tbl} WHERE status = 'completed'")
        except ImportError:
            pass

        try:
            from payments.models import Payment
            _add('pay_total', f"COALESCE(SUM(amount), 0) FROM {Payment._meta.db_table}")
        except ImportError:
            pass

        if SALES_TASK_MANAGER_AVAILABLE:
            try:
                from sales_task_manager.models import SalesTarget, TargetCycle
                cyc_tbl = TargetCycle._meta.db_table
                tgt_tbl = SalesTarget._meta.db_table
                _add('cycle_count', f"COUNT(*) FROM {cyc_tbl} WHERE status = 'ACTIVE'")
                _add('t_target', f"COALESCE(SUM(t.target_amount), 0) FROM {tgt_tbl} t JOIN {cyc_tbl} c ON t.cycle_id = c.id WHERE c.status = 'ACTIVE'")
                _add('t_achieved', f"COALESCE(SUM(t.achieved_amount), 0) FROM {tgt_tbl} t JOIN {cyc_tbl} c ON t.cycle_id = c.id WHERE c.status = 'ACTIVE'")
            except ImportError:
                pass

        proj_tbl = Project._meta.db_table
        item_tbl = WorkItem._meta.db_table
        stat_tbl = WorkItemStatus._meta.db_table
        _add('proj_total', f"COUNT(*) FROM {proj_tbl}")
        _add('proj_active', f"COUNT(*) FROM {proj_tbl} WHERE is_active")
        _add('item_total', f"COUNT(*) FROM {item_tbl}")
        _add('item_todo', f"COUNT(*) FROM {item_tbl} i JOIN {stat_tbl} s ON i.status_id = s.id WHERE s.category = 'todo'")
        _add('item_in_progress', f"COUNT(*) FROM {item_tbl} i JOIN {stat_tbl} s ON i.status_id = s.id WHERE s.category = 'in_progress'")
        _add('item_done', f"COUNT(*) FROM {item_tbl} i JOIN {stat_tbl} s ON i.status_id = s.id WHERE s.category = 'done'")
        _add('item_blocked', f"COUNT(*) FROM {item_tbl} i JOIN {stat_tbl} s ON i.status_id = s.id WHERE s.category = 'blocked'")
        _add('item_overdue', f"COUNT(*) FROM {item_tbl} i JOIN {stat_tbl} s ON i.status_id = s.id WHERE i.due_date < %s AND s.category IN ('todo', 'in_progress')", [now])
        _add('item_unassigned', f"COUNT(*) FROM {item_tbl} WHERE assignee_id IS NULL")

        try:
            from inventory.models import InventoryItem
            invitem_tbl = InventoryItem._meta.db_table
            _add('invitem_total', f"COUNT(*) FROM {invitem_tbl}")
            _add('invitem_active', f"COUNT(*) FROM {invitem_tbl} WHERE status = 'ACTIVE'")
        except ImportError:
            pass

        try:
            from event_calendar.models import CalendarTodo
            _add('cal_overdue', f"COUNT(*) FROM {CalendarTodo._meta.db_table} WHERE \"end\" < %s AND status IS NULL", [now])
        except ImportError:
            pass

        with db_connection.cursor() as cur:
            select_sql = "SELECT " + ", ".join(
                f"({sql}) AS {key}" for key, sql in agg_parts
            )
            cur.execute(select_sql, agg_params)
            row = cur.fetchone()
        rollup = {key: row[i] for i, (key, _) in enumerate(agg_parts)}

        # ── People & HR ────────────────────────────────────────────────
        total_users = rollup.get('user_total', 0)
        active_users = rollup.get('user_active', 0)

        people = {
            'total_users': total_users,
            'active_users': active_users,
            'departments': rollup.get('dept_count', 0),
        }
        hr = {}
        if 'emp_total' in rollup:
            hr['employees_total'] = rollup['emp_total']
            hr['employees_active'] = rollup['emp_active']
            hr['employees_onboarding'] = rollup['emp_onboarding']
            hr['employees_notice'] = rollup['emp_notice']
            hr['leave_pending'] = rollup['leave_pending']
            hr['attendance_today'] = rollup['attendance_today']
            hr['payroll_total'] = float(rollup['payroll_total'])
        people['hr'] = hr

        # ── Contacts ───────────────────────────────────────────────────
        contacts = {'total': 0, 'by_status': [], 'new_30_days': 0}
        if 'contact_total' in rollup:
            from contacts.models import Contact
            cq = Contact.objects.all()
            contacts['total'] = rollup['contact_total']
            contacts['new_30_days'] = rollup['contact_new']
            contacts['by_status'] = list(cq.values('status').annotate(
                count=Count('id')
            ).order_by('-count'))

        # ── CRM ────────────────────────────────────────────────────────
        crm = {
            'deals': 0, 'pipeline_value': 0, 'won_deals': 0,
            'won_value': 0, 'stages': [],
        }
        if 'crm_deals' in rollup:
            from crm.models import CRM, Stage
            deals = CRM.objects.select_related('stage')
            crm['deals'] = rollup['crm_deals']
            crm['pipeline_value'] = float(rollup['crm_pipeline'])
            crm['won_deals'] = rollup['crm_won']
            crm['won_value'] = float(rollup['crm_won_value'])
            stage_rows = deals.values(
                'stage__name', 'stage__slug', 'stage__color'
            ).annotate(
                count=Count('id'),
                value=Sum('value'),
            ).order_by('-value')
            crm['stages'] = []
            for row in stage_rows:
                crm['stages'].append({
                    'name': row['stage__name'],
                    'slug': row['stage__slug'],
                    'color': row['stage__color'],
                    'count': row['count'],
                    'value': float(row['value'] or 0),
                })

        # ── Finance: Invoices + Payments ───────────────────────────────
        finance = {
            'invoices_total': 0,
            'invoices_by_status': [],
            'invoiced_value': 0,
            'paid_value': 0,
            'outstanding': 0,
            'revenue_by_month': [],
            'recent_payments': [],
        }
        if 'inv_total' in rollup:
            from invoices.models import Invoice
            invs = Invoice.objects.all()
            finance['invoices_total'] = rollup['inv_total']
            finance['invoices_by_status'] = list(invs.values('status').annotate(
                count=Count('id')
            ).order_by('status'))
            finance['invoiced_value'] = float(rollup['inv_invoiced'])
        if 'pay_total' in rollup:
            from payments.models import Payment
            pay_qs = Payment.objects.all()
            paid = float(rollup['pay_total'])
            finance['paid_value'] = paid
            finance['outstanding'] = max(
                0, finance['invoiced_value'] - paid
            )
            monthly = pay_qs.annotate(
                month=TruncMonth('created_at')
            ).values('month').annotate(total=Sum('amount')).order_by('month')
            by_month = {str(r['month'])[:7]: float(r['total'] or 0) for r in monthly}
            labels = []
            for i in range(5, -1, -1):
                d = now - timezone.timedelta(days=30 * i)
                labels.append(str(d.date())[:7])
            finance['revenue_by_month'] = [
                {'month': m, 'label': m, 'value': by_month.get(m, 0)}
                for m in labels
            ]
            finance['recent_payments'] = [
                {
                    'id': str(p.id),
                    'amount': float(p.amount),
                    'payment_for': p.payment_for,
                    'payment_method': p.payment_method,
                    'time': p.created_at.isoformat(),
                }
                for p in pay_qs.order_by('-created_at')[:5]
            ]

        # ── Sales Targets ──────────────────────────────────────────────
        sales = {
            'target': 0, 'achieved': 0, 'attainment_pct': 0,
            'active_cycles': 0, 'top_achievers': [],
        }
        if 'cycle_count' in rollup:
            from sales_task_manager.models import SalesTarget, TargetCycle
            cycles = TargetCycle.objects.filter(status='ACTIVE')
            targets = SalesTarget.objects.filter(cycle__in=cycles)
            t_val = float(rollup['t_target'])
            a_val = float(rollup['t_achieved'])
            sales = {
                'target': t_val,
                'achieved': a_val,
                'attainment_pct': round(
                    (a_val / t_val * 100) if t_val else 0, 1
                ),
                'active_cycles': rollup['cycle_count'],
                'top_achievers': [
                    {
                        'name': f"{r['assigned_user__first_name'] or ''} {r['assigned_user__last_name'] or ''}".strip() or r['assigned_user__email'],
                        'achieved': float(r['a'] or 0),
                        'target': float(r['t'] or 0),
                    }
                    for r in targets.filter(
                        assignee_type='USER', assigned_user__isnull=False
                    ).values(
                        'assigned_user__first_name',
                        'assigned_user__last_name',
                        'assigned_user__email',
                    ).annotate(
                        t=Sum('target_amount'), a=Sum('achieved_amount')
                    ).order_by('-a')[:5]
                ],
            }

        # ── Projects & Work ────────────────────────────────────────────
        projects = Project.objects.all()
        active_projects = rollup['proj_active']
        items = WorkItem.objects.select_related('status')
        item_stats = {
            'total': rollup['item_total'],
            'todo': rollup['item_todo'],
            'in_progress': rollup['item_in_progress'],
            'done': rollup['item_done'],
            'blocked': rollup['item_blocked'],
            'overdue': rollup['item_overdue'],
            'unassigned': rollup['item_unassigned'],
            'by_type': list(items.values('issue_type').annotate(
                count=Count('id')
            ).order_by('-count')),
        }
        completion_pct = round(
            (item_stats['done'] / item_stats['total'] * 100)
            if item_stats['total'] else 0, 1
        )
        project_data = {
            'total': rollup['proj_total'],
            'active': active_projects,
            'work_items': item_stats,
            'completion_pct': completion_pct,
        }

        # ── Inventory ──────────────────────────────────────────────────
        inventory = {
            'items': 0, 'active_items': 0, 'low_stock': 0,
            'stock_valuation': 0,
        }
        if 'invitem_total' in rollup:
            from inventory.models import InventoryItem, StockSummary
            inv_items = InventoryItem.objects.all()
            inventory['items'] = rollup['invitem_total']
            inventory['active_items'] = rollup['invitem_active']
            qty_rows = StockSummary.objects.values('item_id').annotate(
                total=Sum('physical_quantity')
            )
            qty_by_item = {r['item_id']: float(r['total'] or 0) for r in qty_rows}
            valuation = 0.0
            low_stock = 0
            for it in inv_items.filter(status='ACTIVE'):
                qty = qty_by_item.get(it.id, 0)
                if it.cost_price:
                    valuation += qty * float(it.cost_price)
                if it.min_stock_level is not None and qty < float(it.min_stock_level):
                    low_stock += 1
            inventory['low_stock'] = low_stock
            inventory['stock_valuation'] = round(valuation, 2)

        # ── Calendar ───────────────────────────────────────────────────
        calendar = {'upcoming': [], 'overdue': 0}
        if 'cal_overdue' in rollup:
            from event_calendar.models import CalendarTodo
            todos = CalendarTodo.objects.all()
            upcoming = todos.filter(
                end__gte=now
            ).order_by('end')[:6]
            calendar['upcoming'] = [
                {
                    'id': str(t.id),
                    'title': t.title,
                    'end_date': t.end.isoformat() if t.end else None,
                }
                for t in upcoming
            ]
            calendar['overdue'] = rollup['cal_overdue']

        # ── Merged recent activity ─────────────────────────────────────
        recent = []
        try:
            for it in WorkItem.objects.select_related('project', 'assignee', 'status') \
                    .order_by('-created_at')[:6]:
                recent.append({
                    'type': 'work_item',
                    'title': it.title,
                    'meta': f"{it.project.key if it.project else ''} • {it.status.name if it.status else 'No status'}",
                    'time': it.created_at.isoformat(),
                })
        except Exception:
            pass
        try:
            from payments.models import Payment
            pay_recent = Payment.objects.order_by('-created_at')[:6]
            for p in pay_recent:
                recent.append({
                    'type': 'payment',
                    'title': p.payment_for,
                    'meta': f"₹{float(p.amount):,.2f} • {p.payment_method or 'Payment'}",
                    'time': p.created_at.isoformat(),
                })
        except ImportError:
            pass
        try:
            from contacts.models import Contact
            for c in Contact.objects.order_by('-created_at')[:5]:
                recent.append({
                    'type': 'contact',
                    'title': c.name or c.email or c.phone or 'Contact',
                    'meta': c.status or 'Lead',
                    'time': c.created_at.isoformat(),
                })
        except Exception:
            pass
        try:
            from invoices.models import Invoice
            for inv in Invoice.objects.order_by('-created_at')[:5]:
                recent.append({
                    'type': 'invoice',
                    'title': inv.invoice_number or 'Invoice',
                    'meta': f"{inv.client_name} • ₹{float(inv.grand_total or 0):,.2f}",
                    'time': inv.created_at.isoformat(),
                })
        except ImportError:
            pass
        recent.sort(key=lambda x: x.get('time', ''), reverse=True)

        # ── KPIs ───────────────────────────────────────────────────────
        org_name = getattr(request.user, 'organization', None)
        try:
            org_name = request.user.organization.name if request.user.organization else None
        except Exception:
            org_name = None

        kpis = {
            'revenue_total': finance['paid_value'],
            'pipeline_value': crm['pipeline_value'],
            'sales_attainment_pct': sales['attainment_pct'],
            'active_projects': active_projects,
            'headcount': hr.get('employees_active', total_users),
            'open_work_items': item_stats['in_progress'] + item_stats['todo'],
            'invoices_total': finance['invoices_total'],
            'low_stock': inventory['low_stock'],
        }

        return Response({
            'generated_at': now.isoformat(),
            'org': {'name': org_name},
            'kpis': kpis,
            'people': people,
            'contacts': contacts,
            'crm': crm,
            'finance': finance,
            'sales': sales,
            'projects': project_data,
            'inventory': inventory,
            'calendar': calendar,
            'recent_activity': recent[:12],
        })

    # ═══════════════════════════════════════════════════════════════════
    # Phase 3: Global Search
    # ═══════════════════════════════════════════════════════════════════

    @action(detail=False, methods=['get'])
    def global_search(self, request):
        """Postgres full-text search across WorkItems (title, description, key, comments)."""
        query = request.query_params.get('q', '').strip()
        workspace_id = request.query_params.get('workspace')
        project_id = request.query_params.get('project')

        if len(query) < 2:
            return Response({'results': [], 'query': query, 'total': 0})

        from django.contrib.postgres.search import SearchVector, SearchQuery, SearchRank

        vector = SearchVector('title', weight='A') + \
                 SearchVector('key', weight='A') + \
                 SearchVector('description', weight='B')

        search_query = SearchQuery(query)

        items = WorkItem.objects.annotate(
            rank=SearchRank(vector, search_query)
        ).filter(rank__gte=0.1).order_by('-rank').select_related(
            'project', 'status', 'assignee'
        )[:25]

        if project_id:
            items = items.filter(project_id=project_id)
        elif workspace_id:
            items = items.filter(project__workspace_id=workspace_id)

        results = []
        for item in items:
            results.append({
                'id': str(item.id),
                'key': item.key,
                'title': item.title,
                'issue_type': item.issue_type,
                'status': item.status.name if item.status else None,
                'status_color': item.status.color if item.status else None,
                'priority': item.priority,
                'project_name': item.project.name,
                'project_key': item.project.key,
                'assignee_name': item.assignee.get_full_name() if item.assignee else None,
                'url': f"/work/{item.project.workspace_id}/{item.project_id}/board?item={item.id}",
                'rank': float(item.rank),
            })

        # Group by project for the frontend
        from collections import OrderedDict
        grouped = OrderedDict()
        for r in results:
            pk = r['project_key']
            if pk not in grouped:
                grouped[pk] = {
                    'project_key': pk,
                    'project_name': r['project_name'],
                    'items': [],
                }
            grouped[pk]['items'].append(r)

        return Response({
            'query': query,
            'total': len(results),
            'results': results,
            'grouped': list(grouped.values()),
        })

    @action(detail=False, methods=['get'])
    def cumulative_flow(self, request):
        project_id = request.query_params.get('project')
        days = int(request.query_params.get('days', 30))
        if not project_id:
            return Response({'error': 'project query param is required'}, status=400)
        from project_management.services.progress_engine import ProgressEngine
        data = ProgressEngine.get_cumulative_flow(project_id, days)
        return Response(data)

    @action(detail=False, methods=['get'])
    def my_dashboard_detail(self, request):
        my_items = WorkItem.objects.filter(assignee=request.user).select_related(
            'project', 'status', 'sprint'
        )
        overdue = my_items.filter(
            due_date__lt=timezone.now(),
            status__category__in=['todo', 'in_progress']
        )
        due_soon = my_items.filter(
            due_date__gte=timezone.now(),
            due_date__lte=timezone.now() + timezone.timedelta(days=3),
            status__category__in=['todo', 'in_progress']
        )
        active_sprint_items = my_items.filter(sprint__status='ACTIVE')
        return Response({
            'total_assigned': my_items.count(),
            'done': my_items.filter(status__category='done').count(),
            'in_progress': my_items.filter(status__category='in_progress').count(),
            'todo': my_items.filter(status__category='todo').count(),
            'blocked': my_items.filter(status__category='blocked').count(),
            'overdue_count': overdue.count(),
            'overdue_items': list(overdue.values('id', 'key', 'title', 'project__name', 'due_date')[:10]),
            'due_soon_count': due_soon.count(),
            'due_soon_items': list(due_soon.values('id', 'key', 'title', 'project__name', 'due_date')[:10]),
            'active_sprint_items': active_sprint_items.count(),
            'project_distribution': list(my_items.values('project__name').annotate(count=Count('id')).order_by('-count')),
            'items_by_type': list(my_items.values('issue_type').annotate(count=Count('id')).order_by('-count')),
        })

    @action(detail=False, methods=['get'])
    def commission_export(self, request):
        from datetime import datetime
        cycle_id = request.query_params.get('cycle')
        workspace_id = request.query_params.get('workspace')
        if not SALES_TASK_MANAGER_AVAILABLE:
            deals = WorkItem.objects.filter(
                issue_type='DEAL', status__slug='won'
            ).select_related('project', 'assignee')
            if workspace_id:
                deals = deals.filter(project__workspace_id=workspace_id)
            rows = []
            for d in deals:
                cf = d.custom_fields or {}
                rows.append({
                    'deal_key': d.key,
                    'title': d.title,
                    'assigned_to': d.assignee.get_full_name() if d.assignee else None,
                    'value': float(cf.get('deal_value', 0)),
                    'won_date': d.completed_at.isoformat() if d.completed_at else None,
                    'project': d.project.name,
                })
            return Response({
                'export_type': 'DEAL',
                'generated_at': datetime.now().isoformat(),
                'rows': rows,
                'total_value': sum(r['value'] for r in rows),
                'count': len(rows),
            })
        try:
            from sales_task_manager.models import TargetCycle, SalesTarget
        except ImportError:
            return Response({'error': 'sales_task_manager not available'}, status=501)
        cycles = TargetCycle.objects.all()
        if cycle_id:
            cycles = cycles.filter(id=cycle_id)
        if workspace_id:
            cycles = cycles.filter(workspace_id=workspace_id)
        rows = []
        for cycle in cycles:
            targets = SalesTarget.objects.filter(cycle=cycle).select_related('assigned_user')
            for t in targets:
                rows.append({
                    'cycle': cycle.name,
                    'assignee': t.assigned_user.get_full_name() if t.assigned_user else None,
                    'target_type': t.assignee_type,
                    'target_amount': float(t.target_amount),
                    'achieved_amount': float(t.achieved_amount),
                    'attainment_pct': round(float(t.achieved_amount) / float(t.target_amount) * 100, 1) if t.target_amount else 0,
                    'status': t.status,
                })
        return Response({
            'export_type': 'TARGET',
            'generated_at': datetime.now().isoformat(),
            'cycles': list(cycles.values('id', 'name', 'code')),
            'rows': rows,
            'total_target': sum(r['target_amount'] for r in rows),
            'total_achieved': sum(r['achieved_amount'] for r in rows),
            'count': len(rows),
        })

    @action(detail=False, methods=['get'])
    def workload(self, request):
        """Cross-project workload & utilization for all users."""
        from project_management.services.resource_engine import get_workload
        data = get_workload(
            workspace_id=request.query_params.get('workspace'),
            department_id=request.query_params.get('department'),
        )
        return Response(data)

    @action(detail=True, methods=['get'])
    def project_workload(self, request, pk=None):
        """Workload for all members of a specific project."""
        from project_management.services.resource_engine import get_project_workload
        data = get_project_workload(pk)
        return Response(data)

    @action(detail=False, methods=['get'])
    def roadmap(self, request):
        """Cross-project roadmap with epics, milestones, and dependencies."""
        from project_management.services.roadmap_engine import get_roadmap_data
        workspace = request.query_params.get('workspace')
        project_ids = request.query_params.getlist('project')
        if not workspace:
            return Response({"error": "workspace parameter is required"}, status=400)
        data = get_roadmap_data(
            workspace_id=workspace,
            project_ids=project_ids if project_ids else None,
            user=request.user,
        )
        return Response(data)

    @action(detail=False, methods=['get'])
    def portfolio(self, request):
        """Cross-project portfolio health, status, and progress."""
        from project_management.services.portfolio_engine import get_portfolio_data
        workspace = request.query_params.get('workspace')
        if not workspace:
            return Response({"error": "workspace parameter is required"}, status=400)
        data = get_portfolio_data(workspace)
        return Response(data)

    @action(detail=False, methods=['get'])
    def time_in_status(self, request):
        """Time-in-status analytics for a project."""
        from project_management.services.analytics_engine import get_time_in_status
        project = request.query_params.get('project')
        days = int(request.query_params.get('days', 90))
        if not project:
            return Response({"error": "project parameter is required"}, status=400)
        data = get_time_in_status(project_id=project, days=days)
        return Response(data)

    @action(detail=False, methods=['get'])
    def cycle_time(self, request):
        """Cycle time and lead time analytics for a project."""
        from project_management.services.analytics_engine import get_cycle_time
        project = request.query_params.get('project')
        days = int(request.query_params.get('days', 90))
        if not project:
            return Response({"error": "project parameter is required"}, status=400)
        data = get_cycle_time(project_id=project, days=days)
        return Response(data)

    @action(detail=False, methods=['get'])
    def capacity_plan(self, request):
        """Automated capacity planning suggestions for a sprint."""
        from project_management.services.capacity_engine import get_capacity_plan
        project = request.query_params.get('project')
        sprint = request.query_params.get('sprint')
        if not project:
            return Response({"error": "project parameter is required"}, status=400)
        data = get_capacity_plan(project_id=project, sprint_id=sprint)
        return Response(data)


# ============================================================================
# RECURRING TASK CONFIG
# ============================================================================

class RecurringTaskConfigViewSet(viewsets.ModelViewSet):
    queryset = RecurringTaskConfig.objects.all()
    serializer_class = RecurringTaskConfigSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['work_item', 'frequency']


# ============================================================================
# APPROVAL WORKFLOWS
# ============================================================================

class ApprovalWorkflowViewSet(viewsets.ModelViewSet):
    queryset = ApprovalWorkflow.objects.all().prefetch_related('steps')
    serializer_class = ApprovalWorkflowSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['project', 'is_active', 'issue_type_filter']


class ApprovalRequestViewSet(viewsets.ModelViewSet):
    queryset = ApprovalRequest.objects.all().select_related(
        'work_item', 'workflow', 'requested_by', 'current_step'
    ).prefetch_related('actions__step', 'actions__user')
    serializer_class = ApprovalRequestSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['work_item', 'status', 'workflow']

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        obj = self.get_object()
        step = obj.current_step
        if not step:
            return Response({'error': 'No active step'}, status=400)
        user = request.user
        if step.assignee_user and step.assignee_user != user:
            return Response({'error': 'You are not the assigned approver'}, status=403)
        action = ApprovalAction.objects.create(
            request=obj, step=step, user=user, action='APPROVE',
            comment=request.data.get('comment', ''),
        )
        if obj.workflow.approval_type == 'SINGLE':
            obj.status = 'APPROVED'
            obj.completed_at = timezone.now()
            obj.current_step = None
        elif obj.workflow.approval_type == 'SEQUENTIAL':
            next_step = ApprovalStep.objects.filter(
                workflow=obj.workflow, step_order__gt=step.step_order
            ).first()
            if next_step:
                obj.current_step = next_step
            else:
                obj.status = 'APPROVED'
                obj.completed_at = timezone.now()
                obj.current_step = None
        obj.save()
        return Response(ApprovalRequestSerializer(obj).data)

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        obj = self.get_object()
        obj.status = 'REJECTED'
        obj.completed_at = timezone.now()
        obj.current_step = None
        obj.save()
        ApprovalAction.objects.create(
            request=obj, step=obj.current_step, user=request.user,
            action='REJECT', comment=request.data.get('comment', ''),
        )
        return Response(ApprovalRequestSerializer(obj).data)


# ============================================================================
# CANNED RESPONSES
# ============================================================================

class CannedResponseViewSet(viewsets.ModelViewSet):
    queryset = CannedResponse.objects.all()
    serializer_class = CannedResponseSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['project', 'category', 'is_global']


# ============================================================================
# PROJECT TEMPLATES
# ============================================================================

class ProjectTemplateViewSet(viewsets.ModelViewSet):
    queryset = ProjectTemplate.objects.all()
    serializer_class = ProjectTemplateSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['category', 'is_global', 'workflow_preset']

    @action(detail=True, methods=['post'])
    def create_project(self, request, pk=None):
        template = self.get_object()
        workspace_id = request.data.get('workspace_id')
        name = request.data.get('name')
        key = request.data.get('key')
        if not workspace_id or not name or not key:
            return Response({'error': 'workspace_id, name, and key are required'}, status=400)
        from project_management.services.workflow_engine import create_project_from_template
        project = create_project_from_template(template, workspace_id, name, key, request.user)
        from project_management.serializers import ProjectSerializer
        return Response(ProjectSerializer(project).data, status=201)


class WorkItemTemplateViewSet(viewsets.ModelViewSet):
    queryset = WorkItemTemplate.objects.all()
    serializer_class = WorkItemTemplateSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['issue_type', 'project', 'category', 'is_global', 'project_template']

    @action(detail=True, methods=['post'])
    def create_item(self, request, pk=None):
        """Create a work item from this template."""
        template = self.get_object()
        project_id = request.data.get('project')
        if not project_id:
            return Response({'error': 'project is required'}, status=400)
        try:
            project = Project.objects.get(id=project_id)
        except Project.DoesNotExist:
            return Response({'error': 'Project not found'}, status=404)

        fields = dict(template.template_fields or {})
        title = request.data.get('title') or fields.get('title') or template.name
        description = request.data.get('description') or fields.get('description', '')
        story_points = request.data.get('story_points') or fields.get('story_points')
        priority = request.data.get('priority') or fields.get('priority', 'MEDIUM')

        from project_management.serializers import WorkItemSerializer
        serializer = WorkItemSerializer(data={
            'project': str(project.id),
            'issue_type': template.issue_type,
            'title': title,
            'description': description,
            'story_points': story_points,
            'priority': priority,
            'template_used': str(template.id),
        }, context={'request': request})
        serializer.is_valid(raise_exception=True)
        item = serializer.save()

        if template.checklist_items:
            item.metadata = {**(item.metadata or {}), 'checklist': template.checklist_items}
            item.save(update_fields=['metadata'])

        return Response(serializer.data, status=201)


# ============================================================================
# GITHUB INTEGRATION
# ============================================================================

class GitHubIntegrationViewSet(viewsets.ModelViewSet):
    queryset = GitHubIntegration.objects.all()
    serializer_class = GitHubIntegrationSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['project', 'is_active']

    @action(detail=True, methods=['post'])
    def sync_webhook(self, request, pk=None):
        obj = self.get_object()
        obj.webhook_secret = uuid.uuid4().hex
        obj.save()
        return Response({'webhook_secret': obj.webhook_secret, 'webhook_url': f'/api/work/github/webhook/{obj.id}/'})


class GitHubLinkViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = GitHubLink.objects.all().select_related('integration', 'work_item')
    serializer_class = GitHubLinkSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['work_item', 'integration', 'link_type', 'state']


# ============================================================================
# PRIORITY RULES
# ============================================================================

# ============================================================================
# MULTI-CHANNEL INTAKE
# ============================================================================

class IntakeViewSet(viewsets.ViewSet):
    permission_classes = [permissions.AllowAny]

    @action(detail=False, methods=['post'])
    def email_to_ticket(self, request):
        """Accept inbound email payload and create a ticket."""
        project_id = request.data.get('project_id')
        subject = request.data.get('subject', '')
        body = request.data.get('body', '')
        sender_email = request.data.get('from', '')
        if not project_id or not subject:
            return Response({'error': 'project_id and subject are required'}, status=400)
        try:
            project = Project.objects.get(id=project_id)
        except Project.DoesNotExist:
            return Response({'error': 'Project not found'}, status=404)
        default_status = WorkItemStatus.objects.filter(workflow=project.workflow).first()
        ticket = WorkItem.objects.create(
            project=project,
            title=subject,
            description=body,
            issue_type='TICKET',
            status=default_status if default_status else None,
            created_by=None,
        )
        if sender_email:
            WorkItemActivityLog.objects.create(
                work_item=ticket,
                activity_type='NOTE',
                description=f"From: {sender_email}",
            )
        from project_management.serializers import WorkItemSerializer
        return Response(WorkItemSerializer(ticket).data, status=201)

    @action(detail=False, methods=['post'])
    def web_form(self, request):
        """Public endpoint for embedded web forms to create tickets."""
        project_id = request.data.get('project_id')
        title = request.data.get('title', '')
        description = request.data.get('description', '')
        email = request.data.get('email', '')
        name = request.data.get('name', '')
        if not project_id or not title:
            return Response({'error': 'project_id and title are required'}, status=400)
        try:
            project = Project.objects.get(id=project_id)
        except Project.DoesNotExist:
            return Response({'error': 'Project not found'}, status=404)
        default_status = WorkItemStatus.objects.filter(workflow=project.workflow).first()
        ticket = WorkItem.objects.create(
            project=project,
            title=title,
            description=description,
            issue_type='TICKET',
            status=default_status if default_status else None,
            created_by=None,
        )
        if email:
            WorkItemActivityLog.objects.create(
                work_item=ticket,
                activity_type='NOTE',
                description=f"Submitted by {name or email} ({email})",
            )
        from project_management.serializers import WorkItemSerializer
        return Response(WorkItemSerializer(ticket).data, status=201)


class PriorityRuleViewSet(viewsets.ModelViewSet):
    queryset = PriorityRule.objects.all()
    serializer_class = PriorityRuleSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['project', 'is_active', 'suggested_priority']

    @action(detail=False, methods=['post'])
    def suggest(self, request):
        title = request.data.get('title', '')
        description = request.data.get('description', '')
        project_id = request.data.get('project_id')
        combined = f"{title} {description}".lower()
        rules = PriorityRule.objects.filter(is_active=True)
        if project_id:
            rules = rules.filter(Q(project_id=project_id) | Q(project__isnull=True))
        best = None
        best_priority = 50
        for rule in rules:
            for kw in rule.keywords:
                if kw.lower() in combined:
                    priority_map = {'LOWEST': 1, 'LOW': 2, 'MEDIUM': 3, 'HIGH': 4, 'HIGHEST': 5, 'CRITICAL': 6}
                    p = priority_map.get(rule.suggested_priority, 3)
                    if p > best_priority:
                        best_priority = p
                        best = rule.suggested_priority
        return Response({'suggested_priority': best or 'MEDIUM'})


# ============================================================================
# WEBHOOK DELIVERY LOGS
# ============================================================================

class WebhookDeliveryLogViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = WebhookDeliveryLog.objects.all().select_related('webhook')
    serializer_class = WebhookDeliveryLogSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['webhook', 'event_type', 'success']


# ============================================================================
# TRANSITION RULES (WORKFLOW GATES)
# ============================================================================

class TransitionRuleViewSet(viewsets.ModelViewSet):
    queryset = TransitionRule.objects.all().select_related('workflow', 'from_status', 'to_status')
    serializer_class = TransitionRuleSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['workflow', 'from_status', 'to_status', 'required_role', 'require_approval']


# ============================================================================
# SAVED FILTERS
# ============================================================================

class SavedFilterViewSet(viewsets.ModelViewSet):
    queryset = SavedFilter.objects.all()
    serializer_class = SavedFilterSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['scope', 'workspace', 'project']

    def get_queryset(self):
        return SavedFilter.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)
